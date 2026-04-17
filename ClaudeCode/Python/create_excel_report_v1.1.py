#!/usr/bin/env python3
"""
GoAmrita Ads — Beautiful Excel Campaign Report Generator v1.1
=============================================================
Author: Msir + Claude
Date: 12 April 2026
Account: Made in Heavens (GoAmrita Bhandar)

v1.1 Changes:
- Column P = "ToDo" (what to do), Column Q = "AI Reason" (why)
- Column R = Action dropdown: "Accept / Review Again" (empty = Skip)
- Json subfolder path for data files
- A08 auto-action is file-level (not per-row)

Usage:
    python create_excel_report_v1.1.py
"""

import json
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================
# CONFIGURATION
# ============================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
REPORT_DATE = "12 April 2026"
REPORT_DIR = os.path.join(PROJECT_DIR, "Report", REPORT_DATE)
JSON_DIR = os.path.join(REPORT_DIR, "Json")

SUMMARY_FILE = os.path.join(JSON_DIR, "report_7day_summary.json")
DAILY_FILE = os.path.join(JSON_DIR, "report_7day_daily.json")
OUTPUT_FILE = os.path.join(REPORT_DIR, "GoAmrita_Campaign_Report_7Day.xlsx")

# ============================================
# STYLES
# ============================================
DARK_BLUE = "1B3A5C"
LIGHT_BLUE_ROW = "E8F0FE"
WHITE_ROW = "FFFFFF"
GREEN_BG = "D5F5E3"
BLUE_BG = "D6EAF8"
YELLOW_BG = "FEF9E7"
PINK_BG = "FADBD8"
RED_BG = "F5B7B1"
DARK_GREEN = "1E8449"
DARK_RED = "C0392B"
SUMMARY_BOX_BG = "ECF0F1"
TOTAL_ROW_BG = "D5D8DC"
TODO_BG = "FDF2E9"
REASON_BG = "F4ECF7"
ACTION_BG = "FDEBD0"

THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color="0D1B2A"),
    right=Side(style="thin", color="0D1B2A"),
    top=Side(style="medium", color="0D1B2A"),
    bottom=Side(style="medium", color="0D1B2A"),
)

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=10, color="2C3E50")
BOLD_DATA = Font(name="Calibri", size=10, bold=True, color="2C3E50")
TOTAL_FONT = Font(name="Calibri", size=11, bold=True, color="1B3A5C")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=DARK_BLUE)
SUBTITLE_FONT = Font(name="Calibri", size=12, bold=True, color="5D6D7E")
LABEL_FONT = Font(name="Calibri", size=11, color="2C3E50")
VALUE_FONT = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
TODO_FONT = Font(name="Calibri", size=10, bold=True, color="D35400")
REASON_FONT = Font(name="Calibri", size=9, color="6C3483")
HEALTH_FONT = Font(name="Calibri", size=12)

HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
ALT_FILL = PatternFill(start_color=LIGHT_BLUE_ROW, end_color=LIGHT_BLUE_ROW, fill_type="solid")
WHITE_FILL = PatternFill(start_color=WHITE_ROW, end_color=WHITE_ROW, fill_type="solid")
TOTAL_FILL = PatternFill(start_color=TOTAL_ROW_BG, end_color=TOTAL_ROW_BG, fill_type="solid")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


# ============================================
# LOAD & ENRICH DATA
# ============================================
def load_data():
    with open(SUMMARY_FILE, "r", encoding="utf-8") as f:
        summary = json.load(f)
    with open(DAILY_FILE, "r", encoding="utf-8") as f:
        daily = json.load(f)

    # Campaign meta from daily (latest per campaign)
    meta = {}
    for row in daily:
        cid = str(row.get("campaignId", ""))
        meta[cid] = {
            "status": row.get("campaignStatus", "UNKNOWN"),
            "budget": float(row.get("campaignBudgetAmount", 0)),
        }

    enriched = []
    for row in summary:
        cid = str(row.get("campaignId", ""))
        m = meta.get(cid, {"status": "UNKNOWN", "budget": 0})

        imp = int(row.get("impressions", 0))
        clk = int(row.get("clicks", 0))
        cost = float(row.get("cost", 0))
        orders = int(row.get("purchases7d", 0))
        sales = float(row.get("sales7d", 0))

        ctr = (clk / imp * 100) if imp > 0 else 0
        cpc = (cost / clk) if clk > 0 else 0
        acos = (cost / sales * 100) if sales > 0 else (999 if cost > 0 else 0)
        roas = (sales / cost) if cost > 0 else 0
        profit = sales - cost

        if sales == 0 and cost > 0:
            health = "\U0001f534"   # red
        elif acos <= 30:
            health = "\U0001f7e2"   # green
        elif acos <= 50:
            health = "\U0001f535"   # blue
        elif acos <= 75:
            health = "\U0001f7e1"   # yellow
        elif acos <= 100:
            health = "\U0001fa77"   # pink heart
        else:
            health = "\U0001f534"   # red

        todo, reason = ai_recommendation(acos, cost, clk, orders, sales, imp, m["budget"])

        enriched.append({
            "name": row.get("campaignName", "Unnamed"),
            "id": cid,
            "status": m["status"],
            "budget": m["budget"],
            "imp": imp, "clk": clk, "ctr": ctr, "cpc": cpc,
            "cost": cost, "orders": orders, "sales": sales,
            "acos": acos if acos < 999 else 0,
            "acos_raw": acos,
            "roas": roas, "profit": profit,
            "health": health,
            "todo": todo, "reason": reason,
        })

    enriched.sort(key=lambda x: x["cost"], reverse=True)
    return enriched


def ai_recommendation(acos, cost, clicks, orders, sales, imp, budget):
    """Returns (ToDo, AI Reason) tuple"""

    if clicks == 0 and imp == 0:
        return (
            "Check Targeting & Bids",
            "No impressions or clicks. Campaign may have targeting issues or bids are too low to win auctions."
        )
    if clicks == 0 and imp > 0:
        return (
            "Increase Bids / Improve Ad Copy",
            f"{imp:,} impressions but 0 clicks (CTR=0%). Ad creative or title not attractive enough, or bid too low for good placement."
        )
    if sales == 0 and cost > 500:
        return (
            "Review Keywords & Listing",
            f"Spent Rs.{cost:,.0f} with {clicks} clicks but ZERO orders. Keywords may be irrelevant or product listing (title/images/price) needs improvement."
        )
    if sales == 0 and cost > 0:
        return (
            "Add Negative Keywords",
            f"Spent Rs.{cost:,.0f} with no sales yet. Irrelevant search terms may be consuming budget. Add negative keywords."
        )
    if acos > 200 and cost > 1000:
        return (
            "PAUSE Campaign",
            f"ACOS {acos:.0f}% — massive loss! Spent Rs.{cost:,.0f} for only {orders} orders (Rs.{sales:,.0f} sales). Net loss: Rs.{cost-sales:,.0f}."
        )
    if acos > 100 and cost > 750:
        return (
            "Reduce Budget 50%",
            f"ACOS {acos:.0f}% — losing money. Every Rs.100 spent returns only Rs.{100/acos*100:.0f}. Cut budget while optimizing keywords."
        )
    if acos > 75:
        return (
            "Lower Bids & Add Negatives",
            f"ACOS {acos:.0f}% — near breakeven. Lower bids on low-converting keywords and add negative keywords for wasteful search terms."
        )
    if acos > 50:
        return (
            "Optimize Keywords & Bids",
            f"ACOS {acos:.0f}% — above target. Pause low-performing keywords, reduce bids on expensive terms, focus on converting ones."
        )
    if acos > 30:
        return (
            "Monitor & Fine-tune",
            f"ACOS {acos:.0f}% — acceptable but improvable. Small bid adjustments and keyword refinement can push ACOS below 30%."
        )
    if acos <= 30 and orders >= 5:
        return (
            "Increase Budget 20%",
            f"ACOS {acos:.0f}% with {orders} orders — PROFITABLE! ROAS {sales/cost:.1f}x. This campaign earns money. Scale it up."
        )
    if acos <= 15 and orders > 0:
        return (
            "Scale Aggressively",
            f"ACOS {acos:.0f}% — excellent winner! ROAS {sales/cost:.1f}x. Increase budget significantly to capture maximum sales."
        )
    if acos <= 30 and orders > 0:
        return (
            "Good — Consider Scaling",
            f"ACOS {acos:.0f}% — profitable but low volume ({orders} orders). Increase bids slightly for more visibility and volume."
        )
    return ("Monitor", "Insufficient data for strong recommendation. Keep monitoring performance.")


# ============================================
# SHEET 1: CAMPAIGN REPORT
# ============================================
def create_campaign_sheet(wb, data):
    ws = wb.active
    ws.title = "Campaign Report"

    # --- Title ---
    ws.merge_cells("A1:R1")
    ws["A1"].value = "GoAmrita Bhandar — Sponsored Products Campaign Report"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:R2")
    ws["A2"].value = f"Period: 5 April - 11 April 2026  |  Account: Made in Heavens  |  Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 5

    # --- Headers (Row 4) --- 18 columns now (A-R)
    headers = [
        "#", "Campaign Name", "Status", "Budget (Rs.)", "Impressions",
        "Clicks", "CTR %", "CPC (Rs.)", "Spend (Rs.)", "Orders (7d)",
        "Sales (Rs.)", "ACOS %", "ROAS", "Profit/Loss (Rs.)", "Health",
        "ToDo", "AI Reason", "Action"
    ]
    col_widths = [5, 42, 12, 14, 14, 10, 10, 12, 14, 12, 14, 12, 10, 16, 9, 28, 50, 16]

    hrow = 4
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=hrow, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci - 1]

    # ToDo header special color
    ws.cell(row=hrow, column=16).fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    # AI Reason header special color
    ws.cell(row=hrow, column=17).fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
    # Action header special color
    ws.cell(row=hrow, column=18).fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")

    ws.row_dimensions[hrow].height = 30

    # --- Data Rows ---
    srow = 5
    for i, c in enumerate(data):
        row = srow + i
        ws.row_dimensions[row].height = 32
        row_fill = ALT_FILL if i % 2 == 0 else WHITE_FILL

        vals = [
            i + 1, c["name"][:50], c["status"], c["budget"],
            c["imp"], c["clk"], c["ctr"], c["cpc"],
            c["cost"], c["orders"], c["sales"], c["acos"],
            c["roas"], c["profit"], c["health"],
            c["todo"], c["reason"], ""
        ]

        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = DATA_FONT
            cell.fill = row_fill
            cell.border = THIN_BORDER

            # Alignment
            if ci in (1, 3, 5, 6, 10, 15, 18):
                cell.alignment = CENTER
            elif ci in (2, 16, 17):
                cell.alignment = LEFT_WRAP
            else:
                cell.alignment = RIGHT

            # Number formats
            if ci in (4, 8, 9, 11, 14):
                cell.number_format = '#,##0.00'
            elif ci in (7, 12):
                cell.number_format = '0.00'
            elif ci == 13:
                cell.number_format = '0.00'
            elif ci == 5:
                cell.number_format = '#,##0'

        # --- ACOS color bands ---
        acos_cell = ws.cell(row=row, column=12)
        av = c["acos"]
        if av == 0 and c["cost"] > 0:
            acos_cell.fill = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
            acos_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        elif av <= 30:
            acos_cell.fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
            acos_cell.font = Font(name="Calibri", size=10, bold=True, color=DARK_GREEN)
        elif av <= 50:
            acos_cell.fill = PatternFill(start_color=BLUE_BG, end_color=BLUE_BG, fill_type="solid")
            acos_cell.font = Font(name="Calibri", size=10, bold=True, color="2471A3")
        elif av <= 75:
            acos_cell.fill = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
            acos_cell.font = Font(name="Calibri", size=10, bold=True, color="B7950B")
        elif av <= 100:
            acos_cell.fill = PatternFill(start_color=PINK_BG, end_color=PINK_BG, fill_type="solid")
            acos_cell.font = Font(name="Calibri", size=10, bold=True, color="C0392B")
        else:
            acos_cell.fill = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
            acos_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

        # Profit/Loss color
        pl_cell = ws.cell(row=row, column=14)
        if c["profit"] >= 0:
            pl_cell.font = Font(name="Calibri", size=10, bold=True, color=DARK_GREEN)
        else:
            pl_cell.font = Font(name="Calibri", size=10, bold=True, color=DARK_RED)

        # Status color
        st_cell = ws.cell(row=row, column=3)
        if c["status"] == "ENABLED":
            st_cell.font = Font(name="Calibri", size=10, bold=True, color=DARK_GREEN)
        elif c["status"] == "PAUSED":
            st_cell.font = Font(name="Calibri", size=10, bold=True, color="E67E22")

        # Health
        ws.cell(row=row, column=15).font = HEALTH_FONT

        # ToDo column styling
        todo_cell = ws.cell(row=row, column=16)
        todo_cell.font = TODO_FONT
        todo_cell.fill = PatternFill(start_color=TODO_BG, end_color=TODO_BG, fill_type="solid")
        todo_cell.alignment = LEFT_WRAP

        # AI Reason column styling
        reason_cell = ws.cell(row=row, column=17)
        reason_cell.font = REASON_FONT
        reason_cell.fill = PatternFill(start_color=REASON_BG, end_color=REASON_BG, fill_type="solid")
        reason_cell.alignment = LEFT_WRAP

        # Action column styling
        action_cell = ws.cell(row=row, column=18)
        action_cell.fill = PatternFill(start_color=ACTION_BG, end_color=ACTION_BG, fill_type="solid")
        action_cell.font = Font(name="Calibri", size=10, bold=True, color="27AE60")
        action_cell.alignment = CENTER

    # --- TOTAL ROW ---
    trow = srow + len(data)
    ws.row_dimensions[trow].height = 32

    t = {
        "imp": sum(x["imp"] for x in data),
        "clk": sum(x["clk"] for x in data),
        "cost": sum(x["cost"] for x in data),
        "orders": sum(x["orders"] for x in data),
        "sales": sum(x["sales"] for x in data),
    }
    t["ctr"] = (t["clk"] / t["imp"] * 100) if t["imp"] > 0 else 0
    t["cpc"] = (t["cost"] / t["clk"]) if t["clk"] > 0 else 0
    t["acos"] = (t["cost"] / t["sales"] * 100) if t["sales"] > 0 else 0
    t["roas"] = (t["sales"] / t["cost"]) if t["cost"] > 0 else 0
    t["profit"] = t["sales"] - t["cost"]

    tvals = [
        "", "TOTAL", f"{len(data)} campaigns", "",
        t["imp"], t["clk"], t["ctr"], t["cpc"],
        t["cost"], t["orders"], t["sales"], t["acos"],
        t["roas"], t["profit"], "", "", "", ""
    ]

    for ci, v in enumerate(tvals, 1):
        cell = ws.cell(row=trow, column=ci, value=v)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = HEADER_BORDER
        if ci in (1, 3, 5, 6, 10, 15, 18):
            cell.alignment = CENTER
        elif ci == 2:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment = RIGHT
        if ci in (4, 8, 9, 11, 14):
            cell.number_format = '#,##0.00'
        elif ci in (7, 12, 13):
            cell.number_format = '0.00'
        elif ci == 5:
            cell.number_format = '#,##0'

    # --- A08 NOTICE ROW ---
    nrow = trow + 2
    ws.merge_cells(f"A{nrow}:R{nrow}")
    notice = ws[f"A{nrow}"]
    notice.value = "A08: If this report is not returned within configured hours, system will auto-apply ALL strong recommendations (PAUSE / Budget cuts) for the entire file."
    notice.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    notice.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    notice.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[nrow].height = 28
    for cc in range(1, 19):
        ws.cell(row=nrow, column=cc).fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")

    # --- LEGEND ROW ---
    lrow = nrow + 1
    ws.merge_cells(f"A{lrow}:R{lrow}")
    legend = ws[f"A{lrow}"]
    legend.value = "Action: Accept = Execute ToDo via API  |  Review Again = Deep analysis with more data  |  Empty = Skip (no action taken)"
    legend.font = Font(name="Calibri", size=9, italic=True, color="7F8C8D")
    legend.alignment = Alignment(horizontal="center")

    # --- DROPDOWN: Accept / Review Again (empty = Skip) ---
    dv = DataValidation(
        type="list",
        formula1='"Accept,Review Again"',
        allow_blank=True,
        showDropDown=False
    )
    dv.prompt = "Accept = execute action  |  Empty = Skip"
    dv.promptTitle = "Select Action"
    dv.error = "Please select: Accept or Review Again (or leave empty to Skip)"
    dv.errorTitle = "Invalid"
    dv.add(f"R{srow}:R{trow - 1}")
    ws.add_data_validation(dv)

    # Freeze & Filter
    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A{hrow}:R{trow}"

    return ws


# ============================================
# SHEET 2: SUMMARY DASHBOARD
# ============================================
def create_summary_sheet(wb, data):
    ws = wb.create_sheet("Summary Dashboard")

    t_imp = sum(c["imp"] for c in data)
    t_clk = sum(c["clk"] for c in data)
    t_cost = sum(c["cost"] for c in data)
    t_orders = sum(c["orders"] for c in data)
    t_sales = sum(c["sales"] for c in data)
    t_ctr = (t_clk / t_imp * 100) if t_imp > 0 else 0
    t_acos = (t_cost / t_sales * 100) if t_sales > 0 else 0
    t_roas = (t_sales / t_cost) if t_cost > 0 else 0
    t_profit = t_sales - t_cost
    t_cpc = (t_cost / t_clk) if t_clk > 0 else 0

    active = len([c for c in data if c["cost"] > 0])
    profitable = len([c for c in data if c["profit"] > 0])
    loss = len([c for c in data if c["profit"] < 0])

    for col, w in [(1, 3), (2, 25), (3, 20), (4, 5), (5, 25), (6, 20), (7, 3)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    # Title
    ws.merge_cells("B1:F1")
    ws["B1"].value = "GoAmrita Bhandar — Performance Dashboard"
    ws["B1"].font = TITLE_FONT
    ws["B1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("B2:F2")
    ws["B2"].value = f"7-Day Report: 5 Apr - 11 Apr 2026  |  Generated: {datetime.now().strftime('%d %b %Y')}"
    ws["B2"].font = SUBTITLE_FONT
    ws["B2"].alignment = Alignment(horizontal="center")

    # Overall Metrics
    row = 4
    ws.merge_cells(f"B{row}:F{row}")
    ws[f"B{row}"].value = "OVERALL PERFORMANCE"
    ws[f"B{row}"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws[f"B{row}"].fill = HEADER_FILL
    ws[f"B{row}"].alignment = CENTER
    for cc in range(2, 7):
        ws.cell(row=row, column=cc).fill = HEADER_FILL
        ws.cell(row=row, column=cc).border = HEADER_BORDER

    box_fill = PatternFill(start_color=SUMMARY_BOX_BG, end_color=SUMMARY_BOX_BG, fill_type="solid")
    metrics = [
        ("Impressions", f"{t_imp:,}", "Total Spend", f"Rs.{t_cost:,.2f}"),
        ("Clicks", f"{t_clk:,}", "Total Sales (7d)", f"Rs.{t_sales:,.2f}"),
        ("CTR", f"{t_ctr:.2f}%", "Profit / Loss", f"Rs.{t_profit:,.2f}"),
        ("CPC", f"Rs.{t_cpc:.2f}", "ACOS", f"{t_acos:.2f}%"),
        ("Orders (7d)", f"{t_orders:,}", "ROAS", f"{t_roas:.2f}x"),
        ("Active Campaigns", f"{active}", "Profitable / Loss", f"{profitable} / {loss}"),
    ]
    for i, (l1, v1, l2, v2) in enumerate(metrics):
        r = row + 1 + i
        ws.row_dimensions[r].height = 26
        ws.cell(row=r, column=2, value=l1).font = LABEL_FONT
        ws.cell(row=r, column=3, value=v1).font = VALUE_FONT
        ws.cell(row=r, column=5, value=l2).font = LABEL_FONT
        ws.cell(row=r, column=6, value=v2).font = VALUE_FONT
        for cc in range(2, 7):
            ws.cell(row=r, column=cc).fill = box_fill
            ws.cell(row=r, column=cc).border = THIN_BORDER
            ws.cell(row=r, column=cc).alignment = Alignment(vertical="center")

    # Profit color
    pc = ws.cell(row=row + 3, column=6)
    pc.font = Font(name="Calibri", size=14, bold=True, color=DARK_GREEN if t_profit >= 0 else DARK_RED)

    # Top 5 Performers
    tr = row + 9
    ws.merge_cells(f"B{tr}:F{tr}")
    ws[f"B{tr}"].value = "TOP 5 PERFORMERS (Lowest ACOS, Orders > 0)"
    ws[f"B{tr}"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    green_fill = PatternFill(start_color=DARK_GREEN, end_color=DARK_GREEN, fill_type="solid")
    for cc in range(2, 7):
        ws.cell(row=tr, column=cc).fill = green_fill

    for ci, h in enumerate(["#", "Campaign", "Spend", "Orders", "ACOS %"], 2):
        cell = ws.cell(row=tr + 1, column=ci, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color="2C3E50")
        cell.fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
        cell.border = THIN_BORDER

    best = sorted([c for c in data if c["orders"] > 0 and c["sales"] > 0], key=lambda x: x["acos"])[:5]
    for i, camp in enumerate(best):
        r = tr + 2 + i
        ws.cell(row=r, column=2, value=i + 1).font = DATA_FONT
        ws.cell(row=r, column=3, value=camp["name"][:35]).font = DATA_FONT
        ws.cell(row=r, column=4, value=f"Rs.{camp['cost']:,.0f}").font = DATA_FONT
        ws.cell(row=r, column=5, value=camp["orders"]).font = DATA_FONT
        ws.cell(row=r, column=6, value=f"{camp['acos']:.1f}%").font = Font(name="Calibri", size=10, bold=True, color=DARK_GREEN)
        for cc in range(2, 7):
            ws.cell(row=r, column=cc).border = THIN_BORDER
            ws.cell(row=r, column=cc).fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")

    # Top 5 Wasters
    wr = tr + 9
    ws.merge_cells(f"B{wr}:F{wr}")
    ws[f"B{wr}"].value = "TOP 5 WORST PERFORMERS (Highest Spend, Bad ACOS)"
    ws[f"B{wr}"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    red_fill = PatternFill(start_color=DARK_RED, end_color=DARK_RED, fill_type="solid")
    for cc in range(2, 7):
        ws.cell(row=wr, column=cc).fill = red_fill

    for ci, h in enumerate(["#", "Campaign", "Spend", "Orders", "ACOS %"], 2):
        cell = ws.cell(row=wr + 1, column=ci, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color="2C3E50")
        cell.fill = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
        cell.border = THIN_BORDER

    worst = sorted([c for c in data if c["cost"] > 100], key=lambda x: (-x["acos_raw"], -x["cost"]))[:5]
    for i, camp in enumerate(worst):
        r = wr + 2 + i
        ad = f"{camp['acos']:.1f}%" if camp["sales"] > 0 else "No Sales"
        ws.cell(row=r, column=2, value=i + 1).font = DATA_FONT
        ws.cell(row=r, column=3, value=camp["name"][:35]).font = DATA_FONT
        ws.cell(row=r, column=4, value=f"Rs.{camp['cost']:,.0f}").font = DATA_FONT
        ws.cell(row=r, column=5, value=camp["orders"]).font = DATA_FONT
        ws.cell(row=r, column=6, value=ad).font = Font(name="Calibri", size=10, bold=True, color=DARK_RED)
        for cc in range(2, 7):
            ws.cell(row=r, column=cc).border = THIN_BORDER
            ws.cell(row=r, column=cc).fill = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")

    # ACOS Distribution
    dr = wr + 9
    ws.merge_cells(f"B{dr}:F{dr}")
    ws[f"B{dr}"].value = "ACOS HEALTH DISTRIBUTION"
    ws[f"B{dr}"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    for cc in range(2, 7):
        ws.cell(row=dr, column=cc).fill = HEADER_FILL

    act = [c for c in data if c["cost"] > 0]
    dist = [
        ("ACOS < 30% (Excellent)", len([c for c in act if c["acos"] <= 30 and c["sales"] > 0]), GREEN_BG),
        ("ACOS 30-50% (Good)", len([c for c in act if 30 < c["acos"] <= 50]), BLUE_BG),
        ("ACOS 50-75% (Average)", len([c for c in act if 50 < c["acos"] <= 75]), YELLOW_BG),
        ("ACOS 75-100% (Bad)", len([c for c in act if 75 < c["acos"] <= 100]), PINK_BG),
        ("ACOS > 100% / No Sales", len([c for c in act if c["acos"] > 100 or (c["sales"] == 0 and c["cost"] > 0)]), RED_BG),
    ]
    for i, (label, count, color) in enumerate(dist):
        r = dr + 1 + i
        ws.cell(row=r, column=2, value=label).font = BOLD_DATA
        ws.cell(row=r, column=3, value=f"{count} campaigns").font = VALUE_FONT
        bar = "\u2588" * count + "\u2591" * max(0, 20 - count)
        ws.cell(row=r, column=5, value=bar).font = Font(name="Consolas", size=10, color="2C3E50")
        for cc in range(2, 7):
            ws.cell(row=r, column=cc).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=r, column=cc).border = THIN_BORDER

    # Wasted spend alert
    ar = dr + 8
    zero_spend = sum(c["cost"] for c in data if c["sales"] == 0 and c["cost"] > 0)
    ws.merge_cells(f"B{ar}:F{ar}")
    ws[f"B{ar}"].value = f"ALERT: Rs.{zero_spend:,.2f} spent on campaigns with ZERO SALES in last 7 days"
    ws[f"B{ar}"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws[f"B{ar}"].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    ws[f"B{ar}"].alignment = CENTER
    for cc in range(2, 7):
        ws.cell(row=ar, column=cc).fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")

    return ws


# ============================================
# MAIN
# ============================================
def main():
    print("=" * 70)
    print("GoAmrita Ads — Excel Report Generator v1.1")
    print(f"Date: {datetime.now().strftime('%d %B %Y, %I:%M %p')}")
    print("=" * 70)

    print("\n Loading data from Json folder...")
    data = load_data()
    print(f" Loaded {len(data)} campaigns")

    print(" Creating Excel workbook...")
    wb = Workbook()

    print(" Building Campaign Report sheet (18 columns)...")
    create_campaign_sheet(wb, data)

    print(" Building Summary Dashboard sheet...")
    create_summary_sheet(wb, data)

    print(f"\n Saving: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    print(f" Size: {os.path.getsize(OUTPUT_FILE):,} bytes")

    print("\n" + "=" * 70)
    print(" Excel report v1.1 generated successfully!")
    print(f" {OUTPUT_FILE}")
    print("\n Columns: # | Name | Status | Budget | Imp | Clicks | CTR% |")
    print("          CPC | Spend | Orders | Sales | ACOS% | ROAS |")
    print("          Profit/Loss | Health | ToDo | AI Reason | Action")
    print("\n Action dropdown: Accept / Review Again (empty = Skip)")
    print(" A08 notice: Auto-apply strong recs if file not returned")
    print("=" * 70)


if __name__ == "__main__":
    main()
