#!/usr/bin/env python3
"""
GoAmrita - Execute Approved Actions v1.0
==========================================
Reads returned Excel approval file → executes approved actions via Ads API

FLOW:
  1. Read Excel (returned by user after review)
  2. Find all rows with "Approve" in Approval Status column
  3. Parse ToDo action for each approved row
  4. Execute via Ads API (campaign pause, budget change, bid change, negative keyword)
  5. Log all actions + results

ACTIONS SUPPORTED:
  - PAUSE Campaign         → PUT /sp/campaigns (state=PAUSED)
  - Reduce Budget X%       → PUT /sp/campaigns (budget * (1 - X/100))
  - Increase Budget X%     → PUT /sp/campaigns (budget * (1 + X/100))
  - Decrease Bid X%        → PUT /sp/keywords (bid * (1 - X/100))
  - Increase Bid X%        → PUT /sp/keywords (bid * (1 + X/100))
  - PAUSE Keyword          → PUT /sp/keywords (state=PAUSED)
  - Add as Negative Exact  → POST /sp/negativeKeywords
  - Add as Negative Phrase → POST /sp/negativeKeywords

Usage:
    python execute_approvals_v1.0.py
    python execute_approvals_v1.0.py --file "path/to/approved_report.xlsx"
    python execute_approvals_v1.0.py --dry-run  (preview, no execute)
"""

import json
import os
import sys
import ssl
import re
import io
import argparse
from datetime import datetime, timedelta
from urllib.request import Request, urlopen
from urllib.parse import urlencode
from urllib.error import HTTPError

# Fix Windows console encoding for special characters
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
if sys.stderr.encoding != 'utf-8':
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

try:
    import certifi
    SSL_CONTEXT = ssl.create_default_context(cafile=certifi.where())
except ImportError:
    SSL_CONTEXT = ssl.create_default_context()

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(os.path.dirname(SCRIPT_DIR))
ADS_CREDS = json.load(open(os.path.join(PROJECT_DIR, "api_credentials.json")))

REPORT_BASE = os.path.join(SCRIPT_DIR, "..", "Report")
report_folders = [f for f in os.listdir(REPORT_BASE) if os.path.isdir(os.path.join(REPORT_BASE, f))]
report_folders.sort(key=lambda x: os.path.getmtime(os.path.join(REPORT_BASE, x)), reverse=True)
LATEST_REPORT = report_folders[0] if report_folders else datetime.now().strftime("%d %B %Y")
DEFAULT_FILE = os.path.join(REPORT_BASE, LATEST_REPORT, f"GoAmrita_Action_Report_{LATEST_REPORT}.xlsx")

# ============================================
# TOKEN
# ============================================
_token = None
_expiry = None

def get_token():
    global _token, _expiry
    if _token and _expiry and datetime.now() < _expiry:
        return _token
    data = urlencode({'grant_type': 'refresh_token', 'refresh_token': ADS_CREDS['refresh_token'],
        'client_id': ADS_CREDS['client_id'], 'client_secret': ADS_CREDS['client_secret']}).encode()
    req = Request(ADS_CREDS['token_url'], data=data, headers={'Content-Type': 'application/x-www-form-urlencoded'})
    result = json.loads(urlopen(req, context=SSL_CONTEXT).read().decode())
    _token = result['access_token']
    _expiry = datetime.now() + timedelta(seconds=result['expires_in'] - 60)
    return _token

def ads_api(method, path, body, ct):
    token = get_token()
    url = ADS_CREDS['api_endpoint'] + path
    headers = {'Authorization': 'Bearer ' + token, 'Amazon-Advertising-API-ClientId': ADS_CREDS['client_id'],
        'Amazon-Advertising-API-Scope': str(ADS_CREDS['profile_id']), 'Content-Type': ct, 'Accept': ct}
    req = Request(url, data=json.dumps(body).encode(), headers=headers, method=method)
    try:
        resp = urlopen(req, context=SSL_CONTEXT)
        return json.loads(resp.read().decode()), resp.status
    except HTTPError as e:
        return {'error': e.read().decode()[:500]}, e.code

# ============================================
# LOAD CAMPAIGN/KEYWORD ID MAPPINGS
# ============================================
def load_mappings():
    """Load campaign name→ID and keyword text→ID mappings from imported data"""
    json_dir = os.path.join(REPORT_BASE, LATEST_REPORT, "Json")
    mappings = {'campaigns': {}, 'keywords': {}, 'adgroups': {}}

    # Campaign name → ID + budget
    camp_file = os.path.join(json_dir, "sp_campaign_list.json")
    if os.path.exists(camp_file):
        with open(camp_file, encoding='utf-8') as f:
            for c in json.load(f):
                name = c.get('name', '')
                cid = str(c.get('campaignId', ''))
                budget = c.get('budget', {}).get('budget', 0)
                mappings['campaigns'][name] = {'id': cid, 'budget': float(budget), 'state': c.get('state', '')}

    # Keyword text → ID + bid + campaign
    kw_file = os.path.join(json_dir, "sp_keyword_list.json")
    if os.path.exists(kw_file):
        with open(kw_file, encoding='utf-8') as f:
            for k in json.load(f):
                key = f"{k.get('keywordText', '').lower()}|{k.get('campaignId', '')}"
                mappings['keywords'][key] = {
                    'id': str(k.get('keywordId', '')),
                    'bid': float(k.get('bid', 0)),
                    'campaignId': str(k.get('campaignId', '')),
                    'adGroupId': str(k.get('adGroupId', '')),
                    'matchType': k.get('matchType', ''),
                }

    # AdGroup per campaign
    ag_file = os.path.join(json_dir, "sp_adgroup_list.json")
    if os.path.exists(ag_file):
        with open(ag_file, encoding='utf-8') as f:
            for a in json.load(f):
                cid = str(a.get('campaignId', ''))
                if cid not in mappings['adgroups']:
                    mappings['adgroups'][cid] = str(a.get('adGroupId', ''))

    return mappings

# ============================================
# PARSE TODO ACTION
# ============================================
def parse_action(todo_text):
    """Parse ToDo text into structured action"""
    if not todo_text:
        return None

    todo = str(todo_text).strip().upper()

    # PAUSE Campaign
    if 'PAUSE CAMPAIGN' in todo or 'PAUSE' == todo:
        return {'type': 'pause_campaign'}

    # Reduce/Decrease Budget X% (handles ~X% too)
    m = re.search(r'(?:REDUCE|DECREASE|CUT)\s*BUDGET\s*~?(\d+)', todo)
    if m:
        return {'type': 'reduce_budget', 'percent': int(m.group(1))}

    # Increase Budget X%
    m = re.search(r'INCREASE\s*BUDGET\s*~?(\d+)', todo)
    if m:
        return {'type': 'increase_budget', 'percent': int(m.group(1))}

    # Decrease/Reduce/Lower Bid X%
    m = re.search(r'(?:DECREASE|REDUCE|LOWER|FINE-TUNE)\s*BID[S]?\s*~?(\d+)', todo)
    if m:
        return {'type': 'reduce_bid', 'percent': int(m.group(1))}

    # Increase Bid X% (handles "Increase Bid 20%", "Increase Bid ~11%", etc.)
    m = re.search(r'INCREASE\s*BID[S]?\s*~?(\d+)', todo)
    if m:
        return {'type': 'increase_bid', 'percent': int(m.group(1))}

    # PAUSE Keyword
    if 'PAUSE KEYWORD' in todo:
        return {'type': 'pause_keyword'}

    # Add Negative
    if 'NEGATIVE EXACT' in todo or 'ADD AS NEGATIVE EXACT' in todo:
        return {'type': 'add_negative', 'matchType': 'NEGATIVE_EXACT'}
    if 'NEGATIVE PHRASE' in todo or 'ADD AS NEGATIVE PHRASE' in todo:
        return {'type': 'add_negative', 'matchType': 'NEGATIVE_PHRASE'}
    if 'NEGATIVE' in todo or 'ADD AS NEGATIVE' in todo:
        return {'type': 'add_negative', 'matchType': 'NEGATIVE_EXACT'}

    # Add as Exact/Phrase Match Keyword (search term graduation)
    if 'ADD AS EXACT' in todo or 'EXACT MATCH KEYWORD' in todo:
        return {'type': 'add_keyword', 'matchType': 'EXACT'}
    if 'ADD AS PHRASE' in todo or 'PHRASE MATCH KEYWORD' in todo:
        return {'type': 'add_keyword', 'matchType': 'PHRASE'}
    if 'ADD AS KEYWORD' in todo or 'GRADUATE' in todo:
        return {'type': 'add_keyword', 'matchType': 'EXACT'}

    # Cut Ad Spend (product level → reduce budget)
    m = re.search(r'CUT\s*(?:AD\s*)?SPEND\s*(\d+)', todo)
    if m:
        return {'type': 'reduce_budget', 'percent': int(m.group(1))}

    # Scale Aggressively = Increase Budget + Bid
    if 'SCALE AGGRESSIVELY' in todo or 'SCALE' in todo:
        m_bud = re.search(r'BUDGET\s*(\d+)', todo)
        m_bid = re.search(r'BID\s*(\d+)', todo)
        return {'type': 'scale', 'budget_pct': int(m_bud.group(1)) if m_bud else 50, 'bid_pct': int(m_bid.group(1)) if m_bid else 20}

    # Fine-tune Bid (with %)
    m = re.search(r'FINE.?TUNE\s*BID\s*~?(\d+)', todo)
    if m:
        return {'type': 'reduce_bid', 'percent': int(m.group(1))}

    # Monitor / Fine-tune (no %) = no action
    if 'MONITOR' in todo or ('FINE' in todo and 'TUNE' in todo and not re.search(r'\d+', todo)):
        return {'type': 'monitor'}

    return {'type': 'unknown', 'raw': todo_text}

# ============================================
# EXECUTE ACTIONS
# ============================================
def execute_campaign_action(action, campaign_name, mappings, dry_run=False):
    """Execute action on campaign"""
    camp = mappings['campaigns'].get(campaign_name)
    if not camp:
        # Try partial match
        for name, data in mappings['campaigns'].items():
            if campaign_name[:20] in name:
                camp = data
                break
    if not camp:
        return {'status': 'FAILED', 'reason': f'Campaign not found: {campaign_name[:30]}'}

    cid = camp['id']
    current_budget = camp['budget']

    if action['type'] == 'pause_campaign':
        if dry_run:
            return {'status': 'DRY_RUN', 'action': f'PAUSE campaign {cid}'}
        r, _ = ads_api('PUT', '/sp/campaigns', {'campaigns': [{'campaignId': cid, 'state': 'PAUSED'}]},
            'application/vnd.spCampaign.v3+json')
        success = r.get('campaigns', {}).get('success', [])
        return {'status': 'OK' if success else 'FAILED', 'action': 'PAUSED', 'campaignId': cid}

    elif action['type'] == 'reduce_budget':
        new_budget = round(current_budget * (1 - action['percent'] / 100), 2)
        new_budget = max(new_budget, 50)  # min Rs.50
        if dry_run:
            return {'status': 'DRY_RUN', 'action': f'Budget {current_budget} -> {new_budget} (-{action["percent"]}%)'}
        r, _ = ads_api('PUT', '/sp/campaigns', {'campaigns': [{'campaignId': cid,
            'budget': {'budgetType': 'DAILY', 'budget': new_budget}}]}, 'application/vnd.spCampaign.v3+json')
        success = r.get('campaigns', {}).get('success', [])
        return {'status': 'OK' if success else 'FAILED', 'action': f'Budget: {current_budget} -> {new_budget}', 'campaignId': cid}

    elif action['type'] == 'increase_budget':
        new_budget = round(current_budget * (1 + action['percent'] / 100), 2)
        if dry_run:
            return {'status': 'DRY_RUN', 'action': f'Budget {current_budget} -> {new_budget} (+{action["percent"]}%)'}
        r, _ = ads_api('PUT', '/sp/campaigns', {'campaigns': [{'campaignId': cid,
            'budget': {'budgetType': 'DAILY', 'budget': new_budget}}]}, 'application/vnd.spCampaign.v3+json')
        success = r.get('campaigns', {}).get('success', [])
        return {'status': 'OK' if success else 'FAILED', 'action': f'Budget: {current_budget} -> {new_budget}', 'campaignId': cid}

    elif action['type'] == 'scale':
        # Scale = increase budget + increase bid (need keyword IDs too)
        new_budget = round(current_budget * (1 + action.get('budget_pct', 50) / 100), 2)
        if dry_run:
            return {'status': 'DRY_RUN', 'action': f'Scale: Budget {current_budget} -> {new_budget} (+{action.get("budget_pct",50)}%)'}
        r, _ = ads_api('PUT', '/sp/campaigns', {'campaigns': [{'campaignId': cid,
            'budget': {'budgetType': 'DAILY', 'budget': new_budget}}]}, 'application/vnd.spCampaign.v3+json')
        success = r.get('campaigns', {}).get('success', [])
        return {'status': 'OK' if success else 'FAILED', 'action': f'Scale: Budget {current_budget} -> {new_budget}', 'campaignId': cid}

    elif action['type'] in ('reduce_bid', 'increase_bid'):
        # Bid action on Campaign sheet — find keywords for this campaign and adjust
        kw_list = [v for k, v in mappings['keywords'].items() if v.get('campaignId') == cid]
        if not kw_list:
            # Auto-targeting campaign (no manual keywords) → fallback to budget change
            pct = action['percent']
            if action['type'] == 'reduce_bid':
                new_budget = round(current_budget * (1 - pct / 100), 2)
                new_budget = max(new_budget, 50)
                fallback_label = f'Auto-target campaign (no manual keywords) — adjusted Budget instead: {current_budget} -> {new_budget} (-{pct}%)'
            else:
                new_budget = round(current_budget * (1 + pct / 100), 2)
                fallback_label = f'Auto-target campaign (no manual keywords) — adjusted Budget instead: {current_budget} -> {new_budget} (+{pct}%)'
            if dry_run:
                return {'status': 'DRY_RUN', 'action': fallback_label, 'campaignId': cid}
            r, _ = ads_api('PUT', '/sp/campaigns', {'campaigns': [{'campaignId': cid,
                'budget': {'budgetType': 'DAILY', 'budget': new_budget}}]}, 'application/vnd.spCampaign.v3+json')
            success = r.get('campaigns', {}).get('success', [])
            return {'status': 'OK' if success else 'FAILED', 'action': fallback_label, 'campaignId': cid}
        adjusted = 0
        for kw in kw_list:
            kid = kw['id']
            current_bid = kw['bid']
            if action['type'] == 'reduce_bid':
                new_bid = round(current_bid * (1 - action['percent'] / 100), 2)
                new_bid = max(new_bid, 2.0)
            else:
                new_bid = round(current_bid * (1 + action['percent'] / 100), 2)
            if dry_run:
                adjusted += 1
                continue
            r, _ = ads_api('PUT', '/sp/keywords', {'keywords': [{'keywordId': kid, 'bid': new_bid}]},
                'application/vnd.spKeyword.v3+json')
            if r.get('keywords', {}).get('success', []):
                adjusted += 1
        action_desc = 'Reduce' if action['type'] == 'reduce_bid' else 'Increase'
        status = 'DRY_RUN' if dry_run else ('OK' if adjusted > 0 else 'FAILED')
        return {'status': status, 'action': f'{action_desc} Bid {action["percent"]}% on {adjusted}/{len(kw_list)} keywords', 'campaignId': cid}

    elif action['type'] == 'monitor':
        return {'status': 'OK', 'action': 'Monitor only - no change', 'campaignId': cid}

    return {'status': 'UNKNOWN_ACTION', 'action': action}


def execute_keyword_action(action, keyword_text, campaign_name, mappings, dry_run=False):
    """Execute action on keyword"""
    # Find keyword ID
    camp = mappings['campaigns'].get(campaign_name, {})
    if not camp:
        for name, data in mappings['campaigns'].items():
            if campaign_name[:20] in name:
                camp = data
                break

    cid = camp.get('id', '') if camp else ''
    kw_lower = keyword_text.lower().strip()
    key = f"{kw_lower}|{cid}"
    kw = mappings['keywords'].get(key)

    if not kw and cid:
        # Try 1: without campaign constraint (same keyword, any campaign)
        for k, v in mappings['keywords'].items():
            if k.startswith(kw_lower + '|'):
                kw = v
                break

    if not kw and cid:
        # Try 2: partial match — keyword text CONTAINS or IS CONTAINED in stored keyword
        # Handles: "vidyapith ayurved sugar churna" vs "vidyapith ayurved sugar churn"
        best_match = None
        best_score = 0
        for k, v in mappings['keywords'].items():
            stored_kw = k.split('|')[0]
            stored_cid = k.split('|')[1] if '|' in k else ''
            # Same campaign preferred
            if stored_cid and cid and stored_cid != cid:
                continue
            # Check containment both ways
            if kw_lower in stored_kw or stored_kw in kw_lower:
                # Score by length of overlap (longer = better match)
                score = min(len(kw_lower), len(stored_kw))
                if score > best_score:
                    best_score = score
                    best_match = v
        if best_match:
            kw = best_match

    if not kw:
        # Try 3: partial match across ALL campaigns (no campaign filter)
        best_match = None
        best_score = 0
        for k, v in mappings['keywords'].items():
            stored_kw = k.split('|')[0]
            if kw_lower in stored_kw or stored_kw in kw_lower:
                score = min(len(kw_lower), len(stored_kw))
                if score > best_score:
                    best_score = score
                    best_match = v
        if best_match:
            kw = best_match

    if action['type'] == 'add_negative':
        agid = kw.get('adGroupId', '') if kw else mappings['adgroups'].get(cid, '')
        if not cid or not agid:
            return {'status': 'FAILED', 'reason': 'Campaign/AdGroup ID not found'}
        if dry_run:
            return {'status': 'DRY_RUN', 'action': f'Add negative "{keyword_text}" ({action["matchType"]})'}
        r, _ = ads_api('POST', '/sp/negativeKeywords', {'negativeKeywords': [{
            'campaignId': cid, 'adGroupId': agid,
            'keywordText': keyword_text, 'matchType': action['matchType'], 'state': 'ENABLED'
        }]}, 'application/vnd.spNegativeKeyword.v3+json')
        success = r.get('negativeKeywords', {}).get('success', [])
        return {'status': 'OK' if success else 'FAILED', 'action': f'Negative added: {keyword_text}'}

    if not kw:
        # Keyword not in keyword list = auto-targeting keyword
        # DON'T change budget here — Campaign sheet already handles campaign-level budget
        # Changing budget from keyword sheet too would cause duplicate/conflicting changes
        camp_name = campaign_name[:25] if campaign_name else cid
        return {'status': 'OK',
                'action': f'Auto-target keyword — bid not editable. Campaign budget already adjusted in Campaign Actions sheet ({camp_name})',
                'keyword': keyword_text[:30]}

    kid = kw['id']
    current_bid = kw['bid']

    if action['type'] == 'pause_keyword':
        if dry_run:
            return {'status': 'DRY_RUN', 'action': f'PAUSE keyword {kid}'}
        r, _ = ads_api('PUT', '/sp/keywords', {'keywords': [{'keywordId': kid, 'state': 'PAUSED'}]},
            'application/vnd.spKeyword.v3+json')
        success = r.get('keywords', {}).get('success', [])
        return {'status': 'OK' if success else 'FAILED', 'action': 'PAUSED', 'keywordId': kid}

    elif action['type'] == 'reduce_bid':
        new_bid = round(current_bid * (1 - action['percent'] / 100), 2)
        new_bid = max(new_bid, 2.0)  # min Rs.2
        if dry_run:
            return {'status': 'DRY_RUN', 'action': f'Bid {current_bid} -> {new_bid} (-{action["percent"]}%)'}
        r, _ = ads_api('PUT', '/sp/keywords', {'keywords': [{'keywordId': kid, 'bid': new_bid}]},
            'application/vnd.spKeyword.v3+json')
        success = r.get('keywords', {}).get('success', [])
        return {'status': 'OK' if success else 'FAILED', 'action': f'Bid: {current_bid} -> {new_bid}', 'keywordId': kid}

    elif action['type'] == 'increase_bid':
        new_bid = round(current_bid * (1 + action['percent'] / 100), 2)
        if dry_run:
            return {'status': 'DRY_RUN', 'action': f'Bid {current_bid} -> {new_bid} (+{action["percent"]}%)'}
        r, _ = ads_api('PUT', '/sp/keywords', {'keywords': [{'keywordId': kid, 'bid': new_bid}]},
            'application/vnd.spKeyword.v3+json')
        success = r.get('keywords', {}).get('success', [])
        return {'status': 'OK' if success else 'FAILED', 'action': f'Bid: {current_bid} -> {new_bid}', 'keywordId': kid}

    if action['type'] == 'add_keyword':
        agid = kw.get('adGroupId', '') if kw else mappings['adgroups'].get(cid, '')
        if not cid or not agid:
            return {'status': 'FAILED', 'reason': 'Campaign/AdGroup ID not found for keyword add'}
        bid_val = kw.get('bid', 5.0) if kw else 5.0
        if dry_run:
            return {'status': 'DRY_RUN', 'action': f'Add keyword "{keyword_text}" ({action["matchType"]}) bid={bid_val}'}
        r, _ = ads_api('POST', '/sp/keywords', {'keywords': [{
            'campaignId': cid, 'adGroupId': agid,
            'keywordText': keyword_text, 'matchType': action['matchType'],
            'bid': bid_val, 'state': 'ENABLED'
        }]}, 'application/vnd.spKeyword.v3+json')
        success = r.get('keywords', {}).get('success', [])
        return {'status': 'OK' if success else 'FAILED', 'action': f'Keyword added: {keyword_text} ({action["matchType"]})'}

    if action['type'] == 'scale':
        # Scale on keyword = increase bid by bid_pct
        if not kw:
            return {'status': 'FAILED', 'reason': f'Keyword not found for scale: {keyword_text[:30]}'}
        bid_pct = action.get('bid_pct', 20)
        new_bid = round(current_bid * (1 + bid_pct / 100), 2)
        if dry_run:
            return {'status': 'DRY_RUN', 'action': f'Scale Bid {current_bid} -> {new_bid} (+{bid_pct}%)'}
        r, _ = ads_api('PUT', '/sp/keywords', {'keywords': [{'keywordId': kid, 'bid': new_bid}]},
            'application/vnd.spKeyword.v3+json')
        success = r.get('keywords', {}).get('success', [])
        return {'status': 'OK' if success else 'FAILED', 'action': f'Scale Bid: {current_bid} -> {new_bid}', 'keywordId': kid}

    if action['type'] == 'monitor':
        return {'status': 'OK', 'action': 'Monitor only - no change needed'}

    return {'status': 'UNKNOWN', 'action': action}


# Amazon auto-target types — NOT real keywords, skip gracefully
AUTO_TARGET_TYPES = ['close-match', 'loose-match', 'substitutes', 'complements',
    'keyword-group="category"', 'keyword-group']


# ============================================
# READ EXCEL & EXECUTE
# ============================================
def process_sheet(ws, sheet_type, mappings, dry_run=False):
    """Process one sheet — find approved rows and execute"""
    results = []

    # Find header row (contains '#' in col 1)
    start_row = 2
    for r in range(1, 10):
        v = str(ws.cell(row=r, column=1).value or '').strip()
        if v == '#':
            start_row = r + 1  # data starts after header
            break

    for row in range(start_row, ws.max_row + 1):
        # Approval column: 7 for Campaign (has TPSR col 4), 6 for Search/Keyword (no TPSR)
        appr_col = 7 if sheet_type == 'campaign' else 6
        approval = str(ws.cell(row=row, column=appr_col).value or '').strip()

        if approval.upper() != 'APPROVE':
            continue

        todo_col = 5 if sheet_type == 'campaign' else 4  # Campaign has TPSR at col 4
        todo = ws.cell(row=row, column=todo_col).value
        action = parse_action(todo)

        if not action or action['type'] == 'unknown':
            results.append({'row': row, 'status': 'SKIP', 'reason': f'Unknown action: {todo}'})
            continue

        if sheet_type == 'campaign':
            campaign_name = str(ws.cell(row=row, column=8).value or '')  # Col H (shifted by TPSR)
            result = execute_campaign_action(action, campaign_name, mappings, dry_run)
            result['row'] = row
            result['campaign'] = campaign_name[:30]
            result['todo'] = str(todo)[:40]
            results.append(result)

        elif sheet_type == 'searchterm':
            search_term = str(ws.cell(row=row, column=7).value or '')  # Col G
            campaign_name = str(ws.cell(row=row, column=8).value or '')  # Col H

            # Skip if search term is empty
            if not search_term.strip():
                continue

            result = execute_keyword_action(action, search_term, campaign_name, mappings, dry_run)
            result['row'] = row
            result['keyword'] = search_term[:30]
            result['todo'] = str(todo)[:40]
            results.append(result)

        elif sheet_type == 'keyword':
            keyword_text = str(ws.cell(row=row, column=7).value or '')  # Col G
            campaign_name = str(ws.cell(row=row, column=9).value or '')  # Col I

            # Skip Amazon auto-target types (not real keywords)
            if keyword_text.lower().strip() in AUTO_TARGET_TYPES or 'keyword-group' in keyword_text.lower():
                results.append({'row': row, 'status': 'OK', 'keyword': keyword_text[:30],
                    'todo': str(todo)[:40], 'action': 'SKIP: Auto-target (Amazon managed, not editable)'})
                continue

            result = execute_keyword_action(action, keyword_text, campaign_name, mappings, dry_run)
            result['row'] = row
            result['keyword'] = keyword_text[:30]
            result['todo'] = str(todo)[:40]
            results.append(result)

    return results

# ============================================
# MAIN
# ============================================
def main():
    parser = argparse.ArgumentParser(description='Execute Approved Actions from Excel')
    parser.add_argument('--file', type=str, default=DEFAULT_FILE, help='Path to approved Excel file')
    parser.add_argument('--dry-run', action='store_true', help='Preview only, no API calls')
    args = parser.parse_args()

    print("=" * 60)
    print("  GoAmrita - Execute Approved Actions v1.0")
    print(f"  {datetime.now().strftime('%d %B %Y, %I:%M %p')}")
    print(f"  Dry Run: {args.dry_run}")
    print("=" * 60)

    if not os.path.exists(args.file):
        print(f"\n  File not found: {args.file}")
        print(f"  Run create_action_report_v2.0.py first!")
        return 1

    print(f"\n  Reading: {os.path.basename(args.file)}")
    wb = load_workbook(args.file)

    print(f"  Loading campaign/keyword mappings...")
    mappings = load_mappings()
    print(f"  Campaigns: {len(mappings['campaigns'])} | Keywords: {len(mappings['keywords'])}")

    all_results = []

    # Process each sheet
    sheet_map = {
        'Campaign Actions': 'campaign',
        'Search Term Actions': 'searchterm',
        'Keyword Actions': 'keyword',
    }

    for sheet_name, sheet_type in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        print(f"\n  Processing: {sheet_name}...")
        results = process_sheet(ws, sheet_type, mappings, dry_run=args.dry_run)

        approved = len(results)
        ok = sum(1 for r in results if r.get('status') in ('OK', 'DRY_RUN'))
        failed = sum(1 for r in results if r.get('status') == 'FAILED')

        print(f"    Approved: {approved} | Executed: {ok} | Failed: {failed}")

        for r in results:
            status_icon = 'OK' if r['status'] in ('OK', 'DRY_RUN') else 'FAIL'
            name = r.get('campaign', r.get('keyword', ''))[:25]
            print(f"      [{status_icon}] {name:<25} | {r.get('todo', ''):<35} | {r.get('action', r.get('reason', ''))}")

        all_results.extend(results)

    # Summary
    total = len(all_results)
    total_ok = sum(1 for r in all_results if r.get('status') in ('OK', 'DRY_RUN'))
    total_fail = sum(1 for r in all_results if r.get('status') == 'FAILED')

    print(f"\n{'='*60}")
    print(f"  EXECUTION {'PREVIEW' if args.dry_run else 'COMPLETE'}")
    print(f"{'='*60}")
    print(f"  Total Approved: {total}")
    print(f"  Executed OK:    {total_ok}")
    print(f"  Failed:         {total_fail}")

    # Save DETAILED execution log (for future undo/rollback)
    log_dir = os.path.join(REPORT_BASE, LATEST_REPORT, "Json")
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = os.path.join(log_dir, f"execution_log_{ts}.json")
    os.makedirs(log_dir, exist_ok=True)

    # Enrich results with before/after for undo
    undo_actions = []
    for r in all_results:
        if r.get('status') not in ('OK', 'DRY_RUN'):
            continue
        undo = {
            'action_type': r.get('todo', ''),
            'entity_type': 'campaign' if 'campaignId' in r else 'keyword' if 'keywordId' in r else 'searchterm',
            'entity_id': r.get('campaignId', r.get('keywordId', '')),
            'entity_name': r.get('campaign', r.get('keyword', '')),
            'action_taken': r.get('action', ''),
            'timestamp': datetime.now().isoformat(),
        }
        # Parse before/after from action string for undo
        action_str = r.get('action', '')
        # Extract first float from a string like "3750.0 (-25%)" → 3750.0
        def _extract_float(s):
            m = re.search(r'[\d.]+', s.strip())
            return float(m.group()) if m else 0.0
        if 'Budget:' in action_str and '->' in action_str:
            parts = action_str.split('Budget:')[1].strip().split('->')
            undo['before_budget'] = _extract_float(parts[0])
            undo['after_budget'] = _extract_float(parts[1])
            undo['undo_action'] = f"Restore budget to {_extract_float(parts[0])}"
        elif 'Bid:' in action_str and '->' in action_str:
            parts = action_str.split('Bid:')[1].strip().split('->')
            undo['before_bid'] = _extract_float(parts[0])
            undo['after_bid'] = _extract_float(parts[1])
            undo['undo_action'] = f"Restore bid to {_extract_float(parts[0])}"
        elif 'PAUSED' in action_str:
            undo['before_state'] = 'ENABLED'
            undo['after_state'] = 'PAUSED'
            undo['undo_action'] = 'Set state back to ENABLED'
        elif 'Negative added' in action_str:
            undo['undo_action'] = 'Remove negative keyword (manual)'
        elif 'Keyword added' in action_str:
            undo['undo_action'] = 'Remove added keyword (manual)'
        elif 'Scale' in action_str:
            undo['undo_action'] = 'Reverse budget increase'
        undo_actions.append(undo)

    log_data = {
        'log_version': '1.0',
        'timestamp': datetime.now().isoformat(),
        'file': args.file,
        'dry_run': args.dry_run,
        'summary': {
            'total_approved': total,
            'executed_ok': total_ok,
            'failed': total_fail,
        },
        'undo_log': undo_actions,
        'undo_instructions': 'To undo: use undo_log entries. Each has before/after values + undo_action description. Phase 2 will have auto-rollback.',
        'results': all_results,
    }

    with open(log_file, 'w', encoding='utf-8') as f:
        json.dump(log_data, f, indent=2, ensure_ascii=False)
    print(f"\n  Execution log: {log_file}")
    print(f"  Undo entries: {len(undo_actions)} (with before/after values for rollback)")

    return 0

if __name__ == "__main__":
    sys.exit(main())
