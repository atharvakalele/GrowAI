"""
GoAmrita Bhandar / Made in Heavens - Listing Health Report Generator v1.0

Backend = Nuclear Reactor (all detection logic hidden)
Frontend = Sweet Pot (clean, simple, listing staff friendly)

Generates professional Excel report with 3 sheets:
  1. Listing Issues (sorted by priority — worst first)
     + Issue Tracking: NEW / REMINDER (7d) / ESCALATED (14d) / RESOLVED / SUPPRESSED
  2. Product Score Card (0-100 score per ASIN)
  3. Legend & Help

v1.0 Changes (13 Apr 2026):
  - Initial release with 9 issue types, score card, legend
  - Added Issue Tracking + Suppress Repeats system
  - History stored in listing_issues_history.json
  - Same issue not shown daily — only NEW, 7d REMINDER, 14d ESCALATED, RESOLVED

Version: 1.0
Date: 13 April 2026
"""

import json
import os
import sys
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ============================================================================
# CONFIGURATION
# ============================================================================

BASE_DIR = Path(r"C:\Users\Clu\Documents\Amazon Bussiness Grouth Automation AI")
REPORT_BASE = BASE_DIR / "ClaudeCode" / "Report"

ACOS_TARGET = 25.0
SELLER_NAME = "GoAmrita Bhandar / Made in Heavens"
ENTITY_ID = "ENTITY1TVPGA5B1GOJW"

# Separate file from Ad Action Report
combine_reports = False

# --- Issue Tracking Config ---
REMINDER_INTERVAL_DAYS = 7       # Days before re-showing a known issue
ESCALATION_AFTER_DAYS = 14       # Days before escalating an unresolved issue
RESOLVED_IMPROVEMENT_THRESHOLD = 0.20  # 20% improvement = resolved

HISTORY_FILE = BASE_DIR / "ClaudeCode" / "Report" / "listing_issues_history.json"

# ============================================================================
# ISSUE HISTORY — LOAD / SAVE
# ============================================================================

def load_issue_history():
    """Load previous issue tracking history. Creates empty dict on first run."""
    if not HISTORY_FILE.exists():
        print(f"  [INFO] No history file found — first run. Will create: {HISTORY_FILE.name}")
        return {}
    try:
        with open(HISTORY_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        print(f"  [OK] Loaded issue history ({len(data)} ASINs tracked)")
        return data
    except Exception as e:
        print(f"  [WARN] Could not load history: {e} — starting fresh")
        return {}

def save_issue_history(history):
    """Save updated issue tracking history to JSON."""
    HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
    try:
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(history, f, indent=2, ensure_ascii=False)
        print(f"  [OK] Issue history saved: {HISTORY_FILE.name} ({len(history)} ASINs)")
    except Exception as e:
        print(f"  [ERROR] Failed to save history: {e}")

def make_issue_key(issue_type):
    """Convert issue type to a short key for history tracking."""
    key_map = {
        "Low Click Rate - Image or Title Problem": "low_ctr",
        "Clicks But No Sales - Listing Not Converting": "zero_conversion",
        "Buy Box Lost - Ads Not Effective": "buybox_lost",
        "Selling But Losing Money on Every Sale": "high_acos",
        "Below Average Click Rate": "below_avg_ctr",
        "Below Average Conversion Rate": "below_avg_cvr",
        "Tight Profit Margin - Ads May Not Be Sustainable": "tight_margin",
        "Low Organic Rank": "low_organic_rank",
        "No Ad Visibility - Zero Impressions": "zero_impressions",
    }
    return key_map.get(issue_type, issue_type.lower().replace(" ", "_")[:30])

def extract_numeric_value(current_value_str):
    """Extract the first numeric value from a current_value string for comparison."""
    import re
    numbers = re.findall(r"[\d.]+", current_value_str.replace(",", ""))
    if numbers:
        try:
            return float(numbers[0])
        except ValueError:
            return 0.0
    return 0.0

def check_improvement(original_val, current_val, issue_key):
    """Check if current value improved enough vs original to mark resolved.

    For most metrics, 'improvement' means:
      - CTR/CVR went UP (higher is better)
      - ACOS went DOWN (lower is better)
      - BSR went DOWN (lower is better)
    Returns True if improved by >= RESOLVED_IMPROVEMENT_THRESHOLD (20%).
    """
    if original_val == 0:
        return False

    # Issues where LOWER is better
    lower_is_better = {"high_acos", "low_organic_rank"}

    if issue_key in lower_is_better:
        # Improvement = value decreased
        if current_val < original_val:
            improvement = (original_val - current_val) / original_val
            return improvement >= RESOLVED_IMPROVEMENT_THRESHOLD
        return False
    else:
        # Improvement = value increased (CTR, CVR, etc.)
        if current_val > original_val:
            improvement = (current_val - original_val) / original_val
            return improvement >= RESOLVED_IMPROVEMENT_THRESHOLD
        return False

def determine_issue_status(asin, issue_key, current_value, history, today_str):
    """Determine what status an issue should have and whether to show it.

    Returns: (status, show_in_report, updated_entry)
      status: "NEW" | "REMINDER" | "ESCALATED" | "RESOLVED" | "SUPPRESSED"
      show_in_report: True/False
      updated_entry: dict to save back into history
    """
    today = date.fromisoformat(today_str)

    asin_history = history.get(asin, {})
    entry = asin_history.get(issue_key, None)

    if entry is None:
        # Brand new issue
        new_entry = {
            "first_detected": today_str,
            "last_shown": today_str,
            "times_shown": 1,
            "status": "open",
            "original_value": current_value,
            "current_value": current_value,
        }
        return "NEW", True, new_entry

    # Existing issue
    if entry.get("status") == "resolved":
        # Was resolved before — if it recurs, treat as NEW again
        new_entry = {
            "first_detected": today_str,
            "last_shown": today_str,
            "times_shown": 1,
            "status": "open",
            "original_value": current_value,
            "current_value": current_value,
        }
        return "NEW", True, new_entry

    # Status is "open" — check for resolution
    original_val = entry.get("original_value", 0)
    if check_improvement(original_val, current_value, issue_key):
        entry["status"] = "resolved"
        entry["current_value"] = current_value
        entry["last_shown"] = today_str
        entry["times_shown"] = entry.get("times_shown", 0) + 1
        return "RESOLVED", True, entry

    # Not resolved — check timing
    first_detected = date.fromisoformat(entry.get("first_detected", today_str))
    last_shown = date.fromisoformat(entry.get("last_shown", today_str))

    days_since_shown = (today - last_shown).days
    days_open = (today - first_detected).days

    entry["current_value"] = current_value

    # Suppress if recently shown
    if days_since_shown < REMINDER_INTERVAL_DAYS:
        return "SUPPRESSED", False, entry

    # Escalate if open for 14+ days
    if days_open >= ESCALATION_AFTER_DAYS:
        entry["last_shown"] = today_str
        entry["times_shown"] = entry.get("times_shown", 0) + 1
        entry["status"] = "open"
        return "ESCALATED", True, entry

    # Reminder if 7+ days since last shown
    if days_since_shown >= REMINDER_INTERVAL_DAYS:
        entry["last_shown"] = today_str
        entry["times_shown"] = entry.get("times_shown", 0) + 1
        return "REMINDER", True, entry

    # Fallback: suppress
    return "SUPPRESSED", False, entry

# ============================================================================
# AUTO-DETECT LATEST REPORT FOLDER
# ============================================================================

def find_latest_report_folder():
    folders = [f for f in REPORT_BASE.iterdir() if f.is_dir()]
    if not folders:
        print("ERROR: No report folders found in", REPORT_BASE)
        sys.exit(1)
    folders.sort(key=lambda f: f.stat().st_mtime, reverse=True)
    return folders[0]

REPORT_FOLDER = find_latest_report_folder()
JSON_DIR = REPORT_FOLDER / "Json"
REPORT_DATE = REPORT_FOLDER.name
OUTPUT_FILENAME = f"GoAmrita_Listing_Health_Report_{REPORT_DATE}.xlsx"

def get_output_path():
    path = REPORT_FOLDER / OUTPUT_FILENAME
    try:
        if path.exists():
            with open(path, "a"):
                pass
        return path
    except PermissionError:
        ts = datetime.now().strftime("%H%M%S")
        alt = REPORT_FOLDER / f"GoAmrita_Listing_Health_Report_{REPORT_DATE}_{ts}.xlsx"
        print(f"  [WARN] Original file locked. Saving as: {alt.name}")
        return alt

OUTPUT_PATH = get_output_path()

print(f"[INFO] Report folder : {REPORT_FOLDER.name}")
print(f"[INFO] JSON directory : {JSON_DIR}")
print(f"[INFO] Output file    : {OUTPUT_PATH}")

# ============================================================================
# STYLE CONSTANTS (matching Ad Action Report)
# ============================================================================

FONT_TITLE = Font(name="Calibri", bold=True, size=16, color="1F4E79")
FONT_SUBTITLE = Font(name="Calibri", bold=True, size=11, color="4472C4")
FONT_STATS_BAR = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
FONT_HEADER = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
FONT_DATA = Font(name="Calibri", size=10)
FONT_DATA_BOLD = Font(name="Calibri", bold=True, size=10)
FONT_SECTION_WHITE = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
FONT_LABEL = Font(name="Calibri", bold=True, size=10, color="333333")
FONT_VALUE = Font(name="Calibri", size=11, color="000000")
FONT_VALUE_BOLD = Font(name="Calibri", bold=True, size=11)

FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FILL_STATS_BAR = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
FILL_SECTION_HDR = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

FILL_RED = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
FILL_GREEN = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
FILL_BLUE = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")

FILL_LIGHT_RED = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
FILL_LIGHT_YELLOW = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
FILL_LIGHT_GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_LIGHT_BLUE = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

FILL_ROW_ALT = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
FILL_ROW_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Priority fills
FILL_PRIORITY_HIGH = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
FILL_PRIORITY_MED = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
FILL_PRIORITY_LOW = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

# Score fills
FILL_SCORE_GREEN = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
FILL_SCORE_BLUE = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
FILL_SCORE_YELLOW = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
FILL_SCORE_RED = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")

FONT_PRIORITY_HIGH = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
FONT_PRIORITY_MED = Font(name="Calibri", bold=True, size=10, color="333333")
FONT_PRIORITY_LOW = Font(name="Calibri", bold=True, size=10, color="333333")
FONT_SCORE_WHITE = Font(name="Calibri", bold=True, size=11, color="FFFFFF")

# Status fills (Issue Tracking)
FILL_STATUS_NEW = PatternFill(start_color="5DADE2", end_color="5DADE2", fill_type="solid")       # Blue
FILL_STATUS_REMINDER = PatternFill(start_color="F7DC6F", end_color="F7DC6F", fill_type="solid")  # Yellow
FILL_STATUS_ESCALATED = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid") # Red
FILL_STATUS_RESOLVED = PatternFill(start_color="58D68D", end_color="58D68D", fill_type="solid")  # Green

FONT_STATUS_NEW = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
FONT_STATUS_REMINDER = Font(name="Calibri", bold=True, size=10, color="333333")
FONT_STATUS_ESCALATED = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
FONT_STATUS_RESOLVED = Font(name="Calibri", bold=True, size=10, color="FFFFFF")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# ============================================================================
# DATA LOADING
# ============================================================================

def load_json(filename):
    filepath = JSON_DIR / filename
    if not filepath.exists():
        print(f"  [WARN] File not found: {filename}")
        return []
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        tag = f"{len(data)} records" if isinstance(data, list) else f"{len(data)} keys"
        print(f"  [OK] Loaded {filename} ({tag})")
        return data
    except Exception as e:
        print(f"  [ERROR] Failed to load {filename}: {e}")
        return []

print("\n--- Loading data files ---")

# Load issue tracking history
issue_history = load_issue_history()
TODAY_STR = date.today().isoformat()  # e.g. "2026-04-13"

product_daily = load_json("sp_advertisedproduct_daily.json")
campaigns_summary = load_json("sp_campaigns_summary.json")
searchterm_daily = load_json("sp_searchterm_daily.json")
true_profit_data = load_json("true_profit_per_asin.json")
pricing_data = load_json("sp_pricing_data.json")
catalog_data = load_json("sp_catalog_prices.json")
product_ads_list = load_json("sp_product_ads_list.json")
buy_box_data = load_json("buy_box_status.json")

# ============================================================================
# BUILD LOOKUP MAPS
# ============================================================================

print("\n--- Building lookup maps ---")

# ASIN -> SKU mapping from product ads list
asin_sku_map = {}
for pad in product_ads_list:
    asin = pad.get("asin", "")
    sku = pad.get("sku", "")
    if asin and sku:
        asin_sku_map[asin] = sku

# Pricing map (dict keyed by ASIN)
price_map = {}
if isinstance(pricing_data, dict):
    for asin, pdata in pricing_data.items():
        price_map[asin] = {
            "your_price": pdata.get("your_price", 0),
            "buy_box_winner": pdata.get("buy_box_winner", None),
        }

# Buy box map (list format)
buybox_map = {}
if isinstance(buy_box_data, list):
    for bb in buy_box_data:
        asin = bb.get("asin", "")
        if asin:
            buybox_map[asin] = {
                "status": bb.get("buy_box_status", ""),
                "is_ours": bb.get("is_ours", False),
                "our_price": bb.get("our_price", 0),
                "bb_price": bb.get("bb_price", 0),
                "reason": bb.get("reason", ""),
                "sku": bb.get("sku", ""),
            }

# Catalog map (dict keyed by ASIN)
catalog_map = {}
if isinstance(catalog_data, dict):
    for asin, cdata in catalog_data.items():
        catalog_map[asin] = {
            "title": cdata.get("title", ""),
            "brand": cdata.get("brand", ""),
            "bsr": cdata.get("bsr", 0),
            "price": cdata.get("price", 0),
        }

# True profit map (list format)
profit_map = {}
if isinstance(true_profit_data, list):
    for tp in true_profit_data:
        asin = tp.get("asin", "")
        if asin:
            profit_map[asin] = {
                "sku": tp.get("sku", ""),
                "sale_price": tp.get("sale_price", 0),
                "product_cost": tp.get("product_cost", 0),
                "true_profit": tp.get("true_profit", 0),
                "true_profit_pct": tp.get("true_profit_pct", 0),
                "profitable_for_ads": tp.get("profitable_for_ads", False),
                "orders_7d": tp.get("orders_7d", 0),
                "sales_7d": tp.get("sales_7d", 0),
                "ad_spend_7d": tp.get("ad_spend_7d", 0),
                "impressions": tp.get("impressions", 0),
                "clicks": tp.get("clicks", 0),
            }

# ============================================================================
# AGGREGATE PER-ASIN AD METRICS (from daily data)
# ============================================================================

print("\n--- Aggregating per-ASIN ad metrics ---")

asin_metrics = defaultdict(lambda: {
    "impressions": 0, "clicks": 0, "cost": 0,
    "orders": 0, "sales": 0, "skus": set(), "campaigns": set()
})

for row in product_daily:
    asin = row.get("advertisedAsin", "")
    if not asin:
        continue
    m = asin_metrics[asin]
    m["impressions"] += row.get("impressions", 0) or 0
    m["clicks"] += row.get("clicks", 0) or 0
    m["cost"] += row.get("cost", 0) or 0
    m["orders"] += row.get("purchases7d", 0) or 0
    m["sales"] += row.get("sales7d", 0) or 0
    sku = row.get("advertisedSku", "")
    if sku:
        m["skus"].add(sku)
    cname = row.get("campaignName", "")
    if cname:
        m["campaigns"].add(cname)

# Also include ASINs from true_profit that have no ad data
for asin in profit_map:
    if asin not in asin_metrics:
        asin_metrics[asin] = {
            "impressions": 0, "clicks": 0, "cost": 0,
            "orders": 0, "sales": 0, "skus": set(), "campaigns": set()
        }

print(f"  [OK] {len(asin_metrics)} unique ASINs found")

# ============================================================================
# COMPUTE ACCOUNT AVERAGES
# ============================================================================

total_impressions = sum(m["impressions"] for m in asin_metrics.values())
total_clicks = sum(m["clicks"] for m in asin_metrics.values())
total_orders = sum(m["orders"] for m in asin_metrics.values())
total_cost = sum(m["cost"] for m in asin_metrics.values())
total_sales = sum(m["sales"] for m in asin_metrics.values())

account_avg_ctr = (total_clicks / total_impressions * 100) if total_impressions > 0 else 0
account_avg_cvr = (total_orders / total_clicks * 100) if total_clicks > 0 else 0
account_avg_acos = (total_cost / total_sales * 100) if total_sales > 0 else 0

print(f"\n--- Account Averages ---")
print(f"  Avg CTR  : {account_avg_ctr:.2f}%")
print(f"  Avg CVR  : {account_avg_cvr:.2f}%")
print(f"  Avg ACOS : {account_avg_acos:.2f}%")

# ============================================================================
# ISSUE DETECTION ENGINE (Nuclear Reactor)
# ============================================================================

print("\n--- Running issue detection engine ---")

issues = []

for asin, m in asin_metrics.items():
    impressions = m["impressions"]
    clicks = m["clicks"]
    orders = m["orders"]
    cost = m["cost"]
    sales = m["sales"]

    ctr = (clicks / impressions * 100) if impressions > 0 else 0
    cvr = (orders / clicks * 100) if clicks > 0 else 0
    acos = (cost / sales * 100) if sales > 0 else 0

    sku = ""
    if m["skus"]:
        sku = sorted(m["skus"])[0]
    elif asin in profit_map:
        sku = profit_map[asin].get("sku", "")
    elif asin in asin_sku_map:
        sku = asin_sku_map[asin]
    elif asin in buybox_map:
        sku = buybox_map[asin].get("sku", "")

    profit_info = profit_map.get(asin, {})
    profit_pct = profit_info.get("true_profit_pct", 0) or 0

    # --- HIGH PRIORITY ISSUES ---

    # 1. LOW CTR + High Impressions
    if impressions > 1000 and ctr < 0.3:
        issues.append({
            "priority": "HIGH",
            "priority_sort": 1,
            "asin": asin,
            "sku": sku,
            "issue_type": "Low Click Rate - Image or Title Problem",
            "problem": f"Got {impressions:,} impressions but only {ctr:.2f}% clicked. Customers see it but don't click.",
            "suggestion": "Main image may not be attractive. Consider lifestyle images. Check title has key benefits and matches search intent.",
            "current_value": f"CTR: {ctr:.2f}%",
            "benchmark": f"Account avg CTR: {account_avg_ctr:.2f}%",
            "impact": f"High impressions wasted without clicks = ad spend inefficiency",
        })

    # 2. ZERO Conversion + High Clicks
    if clicks > 20 and orders == 0:
        issues.append({
            "priority": "HIGH",
            "priority_sort": 2,
            "asin": asin,
            "sku": sku,
            "issue_type": "Clicks But No Sales - Listing Not Converting",
            "problem": f"{clicks} clicks, 0 orders. Visitors click but don't buy.",
            "suggestion": "Check: price vs competitors, reviews/rating, bullet points, A+ content. Listing page may not convince buyers.",
            "current_value": f"0 orders from {clicks} clicks",
            "benchmark": f"Account avg CVR: {account_avg_cvr:.2f}%",
            "impact": f"Rs.{cost:,.0f} wasted on clicks that don't convert",
        })

    # 3. Buy Box Lost
    bb_info = buybox_map.get(asin, {})
    bb_status = bb_info.get("status", "")
    bb_is_ours = bb_info.get("is_ours", True)

    if bb_status in ("NO_BUYBOX", "LOST") or (bb_status and not bb_is_ours):
        reason = bb_info.get("reason", "")
        issues.append({
            "priority": "HIGH",
            "priority_sort": 3,
            "asin": asin,
            "sku": sku,
            "issue_type": "Buy Box Lost - Ads Not Effective",
            "problem": f"Buy Box status: {bb_status}. {reason}",
            "suggestion": "Without Buy Box, Amazon may not show your ads. Check pricing and seller metrics.",
            "current_value": f"Buy Box: {bb_status}",
            "benchmark": "Must have Buy Box for ads to work",
            "impact": "Ads running without Buy Box = mostly wasted spend",
        })

    # 4. Very High ACOS + Some Sales
    if orders > 0 and acos > 100:
        issues.append({
            "priority": "HIGH",
            "priority_sort": 4,
            "asin": asin,
            "sku": sku,
            "issue_type": "Selling But Losing Money on Every Sale",
            "problem": f"ACOS is {acos:.0f}% — spending more than earning on ads.",
            "suggestion": "Either increase price, reduce COGS, or optimize keywords. Current economics don't support ads.",
            "current_value": f"ACOS: {acos:.0f}% | Cost: Rs.{cost:,.0f} | Sales: Rs.{sales:,.0f}",
            "benchmark": f"Target ACOS: {ACOS_TARGET}%",
            "impact": f"Losing Rs.{cost - sales:,.0f} on ad spend vs revenue",
        })

    # --- MEDIUM PRIORITY ISSUES ---

    # 5. CTR Below Average
    if account_avg_ctr > 0 and ctr < (account_avg_ctr * 0.7) and impressions > 500:
        # Skip if already flagged as HIGH for very low CTR
        already_flagged = any(
            i["asin"] == asin and "Low Click Rate" in i["issue_type"]
            for i in issues
        )
        if not already_flagged:
            issues.append({
                "priority": "MEDIUM",
                "priority_sort": 5,
                "asin": asin,
                "sku": sku,
                "issue_type": "Below Average Click Rate",
                "problem": f"CTR ({ctr:.2f}%) is below 70% of account average ({account_avg_ctr:.2f}%).",
                "suggestion": "Title or image underperforming vs your other products. A/B test main image.",
                "current_value": f"CTR: {ctr:.2f}%",
                "benchmark": f"70% of avg = {account_avg_ctr * 0.7:.2f}%",
                "impact": "Lower clicks = fewer potential sales from same impressions",
            })

    # 6. CVR Below Average
    if account_avg_cvr > 0 and clicks > 10 and cvr < (account_avg_cvr * 0.7) and cvr > 0:
        issues.append({
            "priority": "MEDIUM",
            "priority_sort": 6,
            "asin": asin,
            "sku": sku,
            "issue_type": "Below Average Conversion Rate",
            "problem": f"CVR ({cvr:.2f}%) is below 70% of account average ({account_avg_cvr:.2f}%).",
            "suggestion": "Landing page experience needs improvement. Check reviews, pricing, bullet points.",
            "current_value": f"CVR: {cvr:.2f}% ({orders} orders from {clicks} clicks)",
            "benchmark": f"70% of avg = {account_avg_cvr * 0.7:.2f}%",
            "impact": "Getting clicks but not converting — listing page issue",
        })

    # 7. Tight Profit Margin
    if profit_pct != 0 and profit_pct < 20 and profit_pct > -100 and orders > 0:
        issues.append({
            "priority": "MEDIUM",
            "priority_sort": 7,
            "asin": asin,
            "sku": sku,
            "issue_type": "Tight Profit Margin - Ads May Not Be Sustainable",
            "problem": f"True profit margin is only {profit_pct:.1f}%. Very little room for ad spend.",
            "suggestion": "Consider reducing costs or increasing price if market allows. Review if ads are justified.",
            "current_value": f"Margin: {profit_pct:.1f}%",
            "benchmark": "Healthy margin > 20% for sustainable ads",
            "impact": "Low margin + ad cost = potential net loss per sale",
        })

    # --- LOW PRIORITY ISSUES ---

    # 8. High BSR (low organic rank)
    cat_info = catalog_map.get(asin, {})
    bsr = cat_info.get("bsr", 0) or 0
    if bsr > 50000:
        issues.append({
            "priority": "LOW",
            "priority_sort": 9,
            "asin": asin,
            "sku": sku,
            "issue_type": "Low Organic Rank",
            "problem": f"BSR is {bsr:,} — product has low organic visibility.",
            "suggestion": "Organic visibility low. Focus on reviews and keyword optimization. Ensure 7+ images.",
            "current_value": f"BSR: {bsr:,}",
            "benchmark": "Top products BSR < 10,000",
            "impact": "Depends heavily on ads for sales — expensive long term",
        })

    # 9. No Ad Visibility
    if impressions == 0 and asin in profit_map:
        state = profit_map[asin].get("sku", "")
        issues.append({
            "priority": "LOW",
            "priority_sort": 10,
            "asin": asin,
            "sku": sku,
            "issue_type": "No Ad Visibility - Zero Impressions",
            "problem": "Product is getting zero impressions from ads.",
            "suggestion": "Check if keywords are relevant, bids are competitive, or listing is suppressed.",
            "current_value": "0 impressions",
            "benchmark": "Should have some impressions if ads are active",
            "impact": "No ad traffic = completely invisible to ad shoppers",
        })

# Sort: HIGH first (priority_sort asc), then within same priority by cost desc
priority_order = {"HIGH": 1, "MEDIUM": 2, "LOW": 3}
issues.sort(key=lambda x: (priority_order.get(x["priority"], 9), x["priority_sort"]))

print(f"\n--- Raw Issues Detected: {len(issues)} ---")

# ============================================================================
# ISSUE TRACKING — STATUS DETERMINATION & SUPPRESSION
# ============================================================================

print("\n--- Applying issue tracking (suppress repeats) ---")

filtered_issues = []   # Issues that will appear in the report
suppressed_count = 0
status_counts = {"NEW": 0, "REMINDER": 0, "ESCALATED": 0, "RESOLVED": 0}

for issue in issues:
    asin = issue["asin"]
    issue_key = make_issue_key(issue["issue_type"])
    current_numeric = extract_numeric_value(issue["current_value"])

    status, show, updated_entry = determine_issue_status(
        asin, issue_key, current_numeric, issue_history, TODAY_STR
    )

    # Save updated entry back to history
    if asin not in issue_history:
        issue_history[asin] = {}
    issue_history[asin][issue_key] = updated_entry

    if show:
        # Escalated issues get priority bump: MED -> HIGH
        if status == "ESCALATED" and issue["priority"] == "MEDIUM":
            issue["priority"] = "HIGH"

        issue["status"] = status
        issue["first_found"] = updated_entry.get("first_detected", TODAY_STR)
        issue["days_open"] = (date.today() - date.fromisoformat(
            updated_entry.get("first_detected", TODAY_STR)
        )).days

        filtered_issues.append(issue)
        status_counts[status] = status_counts.get(status, 0) + 1
    else:
        suppressed_count += 1

# Also check: any previously tracked issues that are now RESOLVED
# (issue no longer detected = may have improved enough)
# We handle this by marking history entries whose issues didn't appear this run
for asin, asin_entries in issue_history.items():
    for ik, entry in asin_entries.items():
        if entry.get("status") != "open":
            continue
        # Check if this issue was detected in current run
        was_detected = any(
            i["asin"] == asin and make_issue_key(i["issue_type"]) == ik
            for i in issues
        )
        if not was_detected:
            # Issue no longer detected — mark as resolved
            last_shown = date.fromisoformat(entry.get("last_shown", TODAY_STR))
            days_since = (date.today() - last_shown).days
            if days_since >= REMINDER_INTERVAL_DAYS:
                entry["status"] = "resolved"
                entry["last_shown"] = TODAY_STR
                entry["current_value"] = 0
                # Add a resolved entry to the report
                filtered_issues.append({
                    "priority": "LOW",
                    "priority_sort": 99,
                    "asin": asin,
                    "sku": asin_sku_map.get(asin, ""),
                    "issue_type": ik.replace("_", " ").title(),
                    "problem": "Issue no longer detected — appears resolved.",
                    "suggestion": "Great! Monitor to ensure it stays fixed.",
                    "current_value": "Resolved",
                    "benchmark": "-",
                    "impact": "Positive — issue cleared",
                    "status": "RESOLVED",
                    "first_found": entry.get("first_detected", TODAY_STR),
                    "days_open": (date.today() - date.fromisoformat(
                        entry.get("first_detected", TODAY_STR)
                    )).days,
                })
                status_counts["RESOLVED"] = status_counts.get("RESOLVED", 0) + 1

# Re-sort filtered issues: NEW first, then ESCALATED, REMINDER, RESOLVED
status_sort_order = {"NEW": 0, "ESCALATED": 1, "REMINDER": 2, "RESOLVED": 3}
filtered_issues.sort(key=lambda x: (
    status_sort_order.get(x.get("status", "NEW"), 9),
    priority_order.get(x["priority"], 9),
    x.get("priority_sort", 99),
))

# Use filtered_issues for the report from here on
issues_for_report = filtered_issues

high_count = sum(1 for i in issues_for_report if i["priority"] == "HIGH")
med_count = sum(1 for i in issues_for_report if i["priority"] == "MEDIUM")
low_count = sum(1 for i in issues_for_report if i["priority"] == "LOW")

print(f"\n--- Issues Summary (after tracking) ---")
print(f"  NEW        : {status_counts.get('NEW', 0)}")
print(f"  REMINDER   : {status_counts.get('REMINDER', 0)}")
print(f"  ESCALATED  : {status_counts.get('ESCALATED', 0)}")
print(f"  RESOLVED   : {status_counts.get('RESOLVED', 0)}")
print(f"  SUPPRESSED : {suppressed_count} (not shown)")
print(f"  ---")
print(f"  In Report  : {len(issues_for_report)}")
print(f"  HIGH       : {high_count}")
print(f"  MEDIUM     : {med_count}")
print(f"  LOW        : {low_count}")

# ============================================================================
# PRODUCT SCORE CARD CALCULATION
# ============================================================================

print("\n--- Calculating product scores ---")

scorecards = []

for asin, m in asin_metrics.items():
    impressions = m["impressions"]
    clicks = m["clicks"]
    orders = m["orders"]
    cost = m["cost"]
    sales = m["sales"]

    ctr = (clicks / impressions * 100) if impressions > 0 else 0
    cvr = (orders / clicks * 100) if clicks > 0 else 0
    acos = (cost / sales * 100) if sales > 0 else 0

    sku = ""
    if m["skus"]:
        sku = sorted(m["skus"])[0]
    elif asin in profit_map:
        sku = profit_map[asin].get("sku", "")
    elif asin in asin_sku_map:
        sku = asin_sku_map[asin]

    profit_info = profit_map.get(asin, {})
    profit_pct = profit_info.get("true_profit_pct", 0) or 0
    sale_price = profit_info.get("sale_price", 0)

    # Get price from multiple sources
    price = sale_price
    if not price and asin in price_map:
        price = price_map[asin].get("your_price", 0)
    if not price and asin in catalog_map:
        price = catalog_map[asin].get("price", 0)
    if not price and asin in buybox_map:
        price = buybox_map[asin].get("our_price", 0)

    # Product name
    title = catalog_map.get(asin, {}).get("title", "")
    if not title:
        title = sku  # fallback to SKU if no title

    # --- CTR Score (0-25) ---
    if account_avg_ctr > 0 and impressions > 0:
        if ctr > account_avg_ctr * 1.5:
            ctr_score = 25
        elif ctr > account_avg_ctr:
            ctr_score = 20
        elif ctr > account_avg_ctr * 0.7:
            ctr_score = 10
        else:
            ctr_score = 5
    elif impressions == 0:
        ctr_score = 0
    else:
        ctr_score = 10

    # --- CVR Score (0-25) ---
    if account_avg_cvr > 0 and clicks > 0:
        if cvr > account_avg_cvr * 1.5:
            cvr_score = 25
        elif cvr > account_avg_cvr:
            cvr_score = 20
        elif cvr > account_avg_cvr * 0.7:
            cvr_score = 10
        elif cvr > 0:
            cvr_score = 5
        else:
            cvr_score = 0
    elif clicks == 0:
        cvr_score = 0
    else:
        cvr_score = 10

    # --- Profit Score (0-25) ---
    if profit_pct > 50:
        profit_score = 25
    elif profit_pct > 30:
        profit_score = 20
    elif profit_pct > 15:
        profit_score = 15
    elif profit_pct > 0:
        profit_score = 10
    else:
        profit_score = 0

    # --- Ad Efficiency Score (0-25) ---
    if sales > 0:
        if acos < ACOS_TARGET:
            ad_score = 25
        elif acos < ACOS_TARGET * 1.5:
            ad_score = 20
        elif acos < ACOS_TARGET * 2:
            ad_score = 10
        else:
            ad_score = 5
    elif clicks > 0:
        ad_score = 0  # no sales from clicks
    else:
        ad_score = 0  # no data

    overall = ctr_score + cvr_score + profit_score + ad_score

    # Find top issue
    asin_issues = [i for i in issues if i["asin"] == asin]
    top_issue = asin_issues[0]["issue_type"] if asin_issues else "No issues detected"

    # Quick win
    if asin_issues:
        top = asin_issues[0]
        if "Click Rate" in top["issue_type"] or "Image" in top["issue_type"]:
            quick_win = "Improve main image & title"
        elif "No Sales" in top["issue_type"] or "Not Converting" in top["issue_type"]:
            quick_win = "Review price, ratings & bullet points"
        elif "Buy Box" in top["issue_type"]:
            quick_win = "Fix Buy Box — check pricing"
        elif "Losing Money" in top["issue_type"]:
            quick_win = "Reduce bids or increase price"
        elif "Profit Margin" in top["issue_type"]:
            quick_win = "Review pricing & cost structure"
        elif "Organic Rank" in top["issue_type"]:
            quick_win = "Get more reviews & optimize keywords"
        elif "Zero Impressions" in top["issue_type"]:
            quick_win = "Check if listing is active & keywords set"
        else:
            quick_win = "Review listing quality"
    else:
        quick_win = "Maintain current performance"

    scorecards.append({
        "asin": asin,
        "sku": sku,
        "title": title[:60] if title else "",
        "price": price,
        "overall": overall,
        "ctr_score": ctr_score,
        "cvr_score": cvr_score,
        "profit_score": profit_score,
        "ad_score": ad_score,
        "top_issue": top_issue,
        "quick_win": quick_win,
    })

# Sort by overall score ascending (worst first for visibility)
scorecards.sort(key=lambda x: x["overall"])

print(f"  [OK] {len(scorecards)} products scored")

# ============================================================================
# EXCEL GENERATION
# ============================================================================

print("\n--- Generating Excel report ---")

wb = Workbook()

def apply_cell_style(cell, font=None, fill=None, alignment=None, border=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border

def style_header_row(ws, row_num, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        apply_cell_style(cell, FONT_HEADER, FILL_HEADER, ALIGN_CENTER, THIN_BORDER)

def style_data_row(ws, row_num, col_count, alt=False):
    fill = FILL_ROW_ALT if alt else FILL_ROW_WHITE
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        apply_cell_style(cell, FONT_DATA, fill, ALIGN_LEFT, THIN_BORDER)

# ============================================================================
# SHEET 1: LISTING ISSUES
# ============================================================================

ws1 = wb.active
ws1.title = "Listing Issues"
ws1.sheet_properties.tabColor = "FF4444"

# Row 1: Title
ws1.merge_cells("A1:M1")
title_cell = ws1["A1"]
title_cell.value = "GoAmrita \u2014 Listing Health Report"
apply_cell_style(title_cell, FONT_TITLE, None, ALIGN_LEFT)
ws1.row_dimensions[1].height = 36

# Row 2: Subtitle
ws1.merge_cells("A2:M2")
sub_cell = ws1["A2"]
sub_cell.value = f"Period: 7-Day | Report: {REPORT_DATE} | Account: {SELLER_NAME}"
apply_cell_style(sub_cell, FONT_SUBTITLE, None, ALIGN_LEFT)
ws1.row_dimensions[2].height = 22

# Row 3: Summary stats bar (with status breakdown)
ws1.merge_cells("A3:M3")
stats_cell = ws1["A3"]
stats_cell.value = (
    f"Shown: {len(issues_for_report)}  |  "
    f"NEW: {status_counts.get('NEW', 0)}  |  "
    f"REMINDER: {status_counts.get('REMINDER', 0)}  |  "
    f"ESCALATED: {status_counts.get('ESCALATED', 0)}  |  "
    f"RESOLVED: {status_counts.get('RESOLVED', 0)}  |  "
    f"Suppressed: {suppressed_count}  |  "
    f"ASINs: {len(asin_metrics)}"
)
apply_cell_style(stats_cell, FONT_STATS_BAR, FILL_STATS_BAR, ALIGN_CENTER)
ws1.row_dimensions[3].height = 28

# Row 4: Headers (added Status, First Found, Days Open columns)
ISSUE_HEADERS = [
    "#", "Priority", "Status", "ASIN", "SKU", "Issue Type",
    "Problem Found", "AI Suggestion", "Current Value",
    "Benchmark", "First Found", "Days Open", "Impact on Ads"
]

for col_idx, header in enumerate(ISSUE_HEADERS, 1):
    cell = ws1.cell(row=4, column=col_idx, value=header)
    apply_cell_style(cell, FONT_HEADER, FILL_HEADER, ALIGN_CENTER, THIN_BORDER)

ws1.row_dimensions[4].height = 32

# Row 5+: Data (using filtered issues_for_report)
for idx, issue in enumerate(issues_for_report):
    row = 5 + idx
    alt = idx % 2 == 1

    status = issue.get("status", "NEW")
    first_found = issue.get("first_found", TODAY_STR)
    days_open = issue.get("days_open", 0)

    # Status display text
    status_display = status
    if status == "RESOLVED":
        status_display = "RESOLVED \u2705"

    ws1.cell(row=row, column=1, value=idx + 1)
    ws1.cell(row=row, column=2, value=issue["priority"])
    ws1.cell(row=row, column=3, value=status_display)
    ws1.cell(row=row, column=4, value=issue["asin"])
    ws1.cell(row=row, column=5, value=issue["sku"])
    ws1.cell(row=row, column=6, value=issue["issue_type"])
    ws1.cell(row=row, column=7, value=issue["problem"])
    ws1.cell(row=row, column=8, value=issue["suggestion"])
    ws1.cell(row=row, column=9, value=issue["current_value"])
    ws1.cell(row=row, column=10, value=issue["benchmark"])
    ws1.cell(row=row, column=11, value=first_found)
    ws1.cell(row=row, column=12, value=f"{days_open}d" if days_open > 0 else "Today")
    ws1.cell(row=row, column=13, value=issue["impact"])

    # Style all cells
    row_fill = FILL_ROW_ALT if alt else FILL_ROW_WHITE
    for col in range(1, 14):
        cell = ws1.cell(row=row, column=col)
        apply_cell_style(cell, FONT_DATA, row_fill, ALIGN_LEFT, THIN_BORDER)

    # Priority column special styling (col 2)
    priority_cell = ws1.cell(row=row, column=2)
    priority_cell.alignment = ALIGN_CENTER
    if issue["priority"] == "HIGH":
        apply_cell_style(priority_cell, FONT_PRIORITY_HIGH, FILL_PRIORITY_HIGH, ALIGN_CENTER, THIN_BORDER)
    elif issue["priority"] == "MEDIUM":
        apply_cell_style(priority_cell, FONT_PRIORITY_MED, FILL_PRIORITY_MED, ALIGN_CENTER, THIN_BORDER)
    else:
        apply_cell_style(priority_cell, FONT_PRIORITY_LOW, FILL_PRIORITY_LOW, ALIGN_CENTER, THIN_BORDER)

    # Status column special styling (col 3)
    status_cell = ws1.cell(row=row, column=3)
    status_cell.alignment = ALIGN_CENTER
    if status == "NEW":
        apply_cell_style(status_cell, FONT_STATUS_NEW, FILL_STATUS_NEW, ALIGN_CENTER, THIN_BORDER)
    elif status == "REMINDER":
        apply_cell_style(status_cell, FONT_STATUS_REMINDER, FILL_STATUS_REMINDER, ALIGN_CENTER, THIN_BORDER)
    elif status == "ESCALATED":
        apply_cell_style(status_cell, FONT_STATUS_ESCALATED, FILL_STATUS_ESCALATED, ALIGN_CENTER, THIN_BORDER)
    elif status == "RESOLVED":
        apply_cell_style(status_cell, FONT_STATUS_RESOLVED, FILL_STATUS_RESOLVED, ALIGN_CENTER, THIN_BORDER)

    # Center columns: #, First Found, Days Open
    ws1.cell(row=row, column=1).alignment = ALIGN_CENTER
    ws1.cell(row=row, column=11).alignment = ALIGN_CENTER
    ws1.cell(row=row, column=12).alignment = ALIGN_CENTER

    ws1.row_dimensions[row].height = 34

# Column widths (updated for 13 columns)
col_widths_issues = [5, 10, 12, 16, 28, 32, 42, 48, 24, 26, 14, 10, 40]
for i, w in enumerate(col_widths_issues, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Freeze panes
ws1.freeze_panes = "E5"

# ============================================================================
# SHEET 2: PRODUCT SCORE CARD
# ============================================================================

ws2 = wb.create_sheet("Product Score Card")
ws2.sheet_properties.tabColor = "2980B9"

# Row 1: Title
ws2.merge_cells("A1:L1")
title_cell2 = ws2["A1"]
title_cell2.value = "GoAmrita \u2014 Product Score Card"
apply_cell_style(title_cell2, FONT_TITLE, None, ALIGN_LEFT)
ws2.row_dimensions[1].height = 36

# Row 2: Subtitle
ws2.merge_cells("A2:L2")
sub_cell2 = ws2["A2"]
sub_cell2.value = f"Period: 7-Day | Report: {REPORT_DATE} | Scoring: CTR + CVR + Profit + Ad Efficiency (each 0-25 = Total 0-100)"
apply_cell_style(sub_cell2, FONT_SUBTITLE, None, ALIGN_LEFT)
ws2.row_dimensions[2].height = 22

# Row 3: Stats bar
green_count = sum(1 for s in scorecards if s["overall"] >= 75)
blue_count = sum(1 for s in scorecards if 50 <= s["overall"] < 75)
yellow_count = sum(1 for s in scorecards if 25 <= s["overall"] < 50)
red_count = sum(1 for s in scorecards if s["overall"] < 25)

ws2.merge_cells("A3:L3")
stats_cell2 = ws2["A3"]
stats_cell2.value = (
    f"Products: {len(scorecards)}  |  "
    f"Green (75-100): {green_count}  |  "
    f"Blue (50-74): {blue_count}  |  "
    f"Yellow (25-49): {yellow_count}  |  "
    f"Red (0-24): {red_count}"
)
apply_cell_style(stats_cell2, FONT_STATS_BAR, FILL_STATS_BAR, ALIGN_CENTER)
ws2.row_dimensions[3].height = 28

# Row 4: Headers
SCORE_HEADERS = [
    "#", "ASIN", "SKU", "Product Name", "Price",
    "Overall Score", "CTR Score", "CVR Score", "Profit Score",
    "Ad Efficiency", "Top Issue", "Quick Win"
]

for col_idx, header in enumerate(SCORE_HEADERS, 1):
    cell = ws2.cell(row=4, column=col_idx, value=header)
    apply_cell_style(cell, FONT_HEADER, FILL_HEADER, ALIGN_CENTER, THIN_BORDER)

ws2.row_dimensions[4].height = 32

# Row 5+: Data
for idx, sc in enumerate(scorecards):
    row = 5 + idx
    alt = idx % 2 == 1

    ws2.cell(row=row, column=1, value=idx + 1)
    ws2.cell(row=row, column=2, value=sc["asin"])
    ws2.cell(row=row, column=3, value=sc["sku"])
    ws2.cell(row=row, column=4, value=sc["title"])
    ws2.cell(row=row, column=5, value=sc["price"] if sc["price"] else "N/A")
    ws2.cell(row=row, column=6, value=sc["overall"])
    ws2.cell(row=row, column=7, value=f'{sc["ctr_score"]}/25')
    ws2.cell(row=row, column=8, value=f'{sc["cvr_score"]}/25')
    ws2.cell(row=row, column=9, value=f'{sc["profit_score"]}/25')
    ws2.cell(row=row, column=10, value=f'{sc["ad_score"]}/25')
    ws2.cell(row=row, column=11, value=sc["top_issue"])
    ws2.cell(row=row, column=12, value=sc["quick_win"])

    # Style all cells
    row_fill = FILL_ROW_ALT if alt else FILL_ROW_WHITE
    for col in range(1, 13):
        cell = ws2.cell(row=row, column=col)
        apply_cell_style(cell, FONT_DATA, row_fill, ALIGN_LEFT, THIN_BORDER)

    # Center columns: #, Price, Scores
    ws2.cell(row=row, column=1).alignment = ALIGN_CENTER
    ws2.cell(row=row, column=5).alignment = ALIGN_CENTER
    for c in range(6, 11):
        ws2.cell(row=row, column=c).alignment = ALIGN_CENTER

    # Overall Score color
    score_cell = ws2.cell(row=row, column=6)
    overall = sc["overall"]
    if overall >= 75:
        apply_cell_style(score_cell, FONT_SCORE_WHITE, FILL_SCORE_GREEN, ALIGN_CENTER, THIN_BORDER)
    elif overall >= 50:
        apply_cell_style(score_cell, FONT_SCORE_WHITE, FILL_SCORE_BLUE, ALIGN_CENTER, THIN_BORDER)
    elif overall >= 25:
        apply_cell_style(score_cell, FONT_DATA_BOLD, FILL_SCORE_YELLOW, ALIGN_CENTER, THIN_BORDER)
    else:
        apply_cell_style(score_cell, FONT_SCORE_WHITE, FILL_SCORE_RED, ALIGN_CENTER, THIN_BORDER)

    ws2.row_dimensions[row].height = 32

# Column widths
col_widths_score = [5, 16, 28, 42, 10, 14, 12, 12, 12, 14, 34, 34]
for i, w in enumerate(col_widths_score, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Freeze panes
ws2.freeze_panes = "C5"

# ============================================================================
# SHEET 3: LEGEND & HELP
# ============================================================================

ws3 = wb.create_sheet("Legend & Help")
ws3.sheet_properties.tabColor = "27AE60"

# Title
ws3.merge_cells("A1:F1")
ws3["A1"].value = "Listing Health Report \u2014 Legend & Help"
apply_cell_style(ws3["A1"], FONT_TITLE, None, ALIGN_LEFT)
ws3.row_dimensions[1].height = 36

# --- Section 1: Priority Colors ---
row = 3
ws3.merge_cells(f"A{row}:F{row}")
ws3[f"A{row}"].value = "Priority Colors"
apply_cell_style(ws3[f"A{row}"], FONT_SECTION_WHITE, FILL_SECTION_HDR, ALIGN_LEFT)

row += 1
legend_priorities = [
    ("HIGH", FILL_PRIORITY_HIGH, FONT_PRIORITY_HIGH, "Immediate attention needed. Fix this ASAP or stop ad spend."),
    ("MEDIUM", FILL_PRIORITY_MED, FONT_PRIORITY_MED, "Should improve. Underperforming vs account average."),
    ("LOW", FILL_PRIORITY_LOW, FONT_PRIORITY_LOW, "Nice to have. Improvement would help long-term health."),
]
for label, fill, font, desc in legend_priorities:
    ws3.cell(row=row, column=1, value=label)
    apply_cell_style(ws3.cell(row=row, column=1), font, fill, ALIGN_CENTER, THIN_BORDER)
    ws3.merge_cells(f"B{row}:F{row}")
    ws3.cell(row=row, column=2, value=desc)
    apply_cell_style(ws3.cell(row=row, column=2), FONT_DATA, None, ALIGN_LEFT, THIN_BORDER)
    ws3.row_dimensions[row].height = 26
    row += 1

# --- Section 2: Score Colors ---
row += 1
ws3.merge_cells(f"A{row}:F{row}")
ws3[f"A{row}"].value = "Score Card Colors (0-100)"
apply_cell_style(ws3[f"A{row}"], FONT_SECTION_WHITE, FILL_SECTION_HDR, ALIGN_LEFT)

row += 1
score_legend = [
    ("75-100 Green", FILL_SCORE_GREEN, FONT_SCORE_WHITE, "Excellent performer. Maintain and scale."),
    ("50-74 Blue", FILL_SCORE_BLUE, FONT_SCORE_WHITE, "Good with room for improvement."),
    ("25-49 Yellow", FILL_SCORE_YELLOW, FONT_DATA_BOLD, "Needs attention. Multiple areas to fix."),
    ("0-24 Red", FILL_SCORE_RED, FONT_SCORE_WHITE, "Critical. Listing has major issues."),
]
for label, fill, font, desc in score_legend:
    ws3.cell(row=row, column=1, value=label)
    apply_cell_style(ws3.cell(row=row, column=1), font, fill, ALIGN_CENTER, THIN_BORDER)
    ws3.merge_cells(f"B{row}:F{row}")
    ws3.cell(row=row, column=2, value=desc)
    apply_cell_style(ws3.cell(row=row, column=2), FONT_DATA, None, ALIGN_LEFT, THIN_BORDER)
    ws3.row_dimensions[row].height = 26
    row += 1

# --- Section 3: Issue Status Colors ---
row += 1
ws3.merge_cells(f"A{row}:F{row}")
ws3[f"A{row}"].value = "Issue Tracking Status"
apply_cell_style(ws3[f"A{row}"], FONT_SECTION_WHITE, FILL_SECTION_HDR, ALIGN_LEFT)

row += 1
status_legend = [
    ("NEW", FILL_STATUS_NEW, FONT_STATUS_NEW,
     "First time this issue is detected. Needs attention."),
    ("REMINDER", FILL_STATUS_REMINDER, FONT_STATUS_REMINDER,
     f"Issue was shown {REMINDER_INTERVAL_DAYS}+ days ago, still not fixed. Showing again."),
    ("ESCALATED", FILL_STATUS_ESCALATED, FONT_STATUS_ESCALATED,
     f"Open for {ESCALATION_AFTER_DAYS}+ days. Priority bumped up. Fix urgently."),
    ("RESOLVED \u2705", FILL_STATUS_RESOLVED, FONT_STATUS_RESOLVED,
     "Issue improved by 20%+ or no longer detected. Good job!"),
]
for label, fill, font, desc in status_legend:
    ws3.cell(row=row, column=1, value=label)
    apply_cell_style(ws3.cell(row=row, column=1), font, fill, ALIGN_CENTER, THIN_BORDER)
    ws3.merge_cells(f"B{row}:F{row}")
    ws3.cell(row=row, column=2, value=desc)
    apply_cell_style(ws3.cell(row=row, column=2), FONT_DATA, None, ALIGN_LEFT, THIN_BORDER)
    ws3.row_dimensions[row].height = 26
    row += 1

# --- Section 4: Score Breakdown ---
row += 1
ws3.merge_cells(f"A{row}:F{row}")
ws3[f"A{row}"].value = "Score Breakdown (Each Component: 0-25)"
apply_cell_style(ws3[f"A{row}"], FONT_SECTION_WHITE, FILL_SECTION_HDR, ALIGN_LEFT)

row += 1
score_breakdown = [
    ("CTR Score (0-25)", "Based on Click-Through Rate vs account average. Higher CTR = better image & title."),
    ("CVR Score (0-25)", "Based on Conversion Rate. Higher CVR = better listing page (price, reviews, content)."),
    ("Profit Score (0-25)", "Based on True Profit Margin. Higher margin = more room for ad spend."),
    ("Ad Efficiency (0-25)", "Based on ACOS vs target. Lower ACOS = more efficient ad spend."),
]
for label, desc in score_breakdown:
    ws3.cell(row=row, column=1, value=label)
    apply_cell_style(ws3.cell(row=row, column=1), FONT_DATA_BOLD, FILL_LIGHT_BLUE, ALIGN_LEFT, THIN_BORDER)
    ws3.merge_cells(f"B{row}:F{row}")
    ws3.cell(row=row, column=2, value=desc)
    apply_cell_style(ws3.cell(row=row, column=2), FONT_DATA, None, ALIGN_LEFT, THIN_BORDER)
    ws3.row_dimensions[row].height = 28
    row += 1

# --- Section 5: Issue Types Explained ---
row += 1
ws3.merge_cells(f"A{row}:F{row}")
ws3[f"A{row}"].value = "Issue Types Explained"
apply_cell_style(ws3[f"A{row}"], FONT_SECTION_WHITE, FILL_SECTION_HDR, ALIGN_LEFT)

row += 1
issue_types = [
    ("Low Click Rate", "Many people see your product but don't click. Usually a main image or title problem."),
    ("Clicks But No Sales", "People click your ad but don't buy. Listing page not convincing — price, reviews, bullets."),
    ("Buy Box Lost", "You don't own the Buy Box. Amazon may not show your ads. Check pricing & seller health."),
    ("Losing Money on Sales", "ACOS > 100%. You spend more on ads than you earn. Unsustainable."),
    ("Below Avg Click Rate", "Your CTR is lower than 70% of your account average. Image/title underperforming."),
    ("Below Avg Conversion", "Your CVR is lower than 70% of account average. Listing page needs work."),
    ("Tight Profit Margin", "Profit margin < 20%. Not much room for ad spend. May need price/cost change."),
    ("Low Organic Rank", "BSR > 50,000. Product relies heavily on ads. Improve organic visibility."),
    ("Zero Impressions", "No ad impressions at all. Listing may be suppressed or keywords not set."),
]
for label, desc in issue_types:
    ws3.cell(row=row, column=1, value=label)
    apply_cell_style(ws3.cell(row=row, column=1), FONT_DATA_BOLD, None, ALIGN_LEFT, THIN_BORDER)
    ws3.merge_cells(f"B{row}:F{row}")
    ws3.cell(row=row, column=2, value=desc)
    apply_cell_style(ws3.cell(row=row, column=2), FONT_DATA, None, ALIGN_LEFT, THIN_BORDER)
    ws3.row_dimensions[row].height = 28
    row += 1

# --- Section 6: Account Averages ---
row += 1
ws3.merge_cells(f"A{row}:F{row}")
ws3[f"A{row}"].value = "Account Averages (Benchmarks)"
apply_cell_style(ws3[f"A{row}"], FONT_SECTION_WHITE, FILL_SECTION_HDR, ALIGN_LEFT)

row += 1
benchmarks = [
    ("Account Avg CTR", f"{account_avg_ctr:.2f}%"),
    ("Account Avg CVR", f"{account_avg_cvr:.2f}%"),
    ("Account Avg ACOS", f"{account_avg_acos:.2f}%"),
    ("ACOS Target", f"{ACOS_TARGET:.0f}%"),
    ("Report Period", "7 Days"),
    ("Data Source", f"Ads API Report: {REPORT_DATE}"),
]
for label, value in benchmarks:
    ws3.cell(row=row, column=1, value=label)
    apply_cell_style(ws3.cell(row=row, column=1), FONT_LABEL, FILL_LIGHT_BLUE, ALIGN_LEFT, THIN_BORDER)
    ws3.cell(row=row, column=2, value=value)
    apply_cell_style(ws3.cell(row=row, column=2), FONT_VALUE_BOLD, None, ALIGN_LEFT, THIN_BORDER)
    ws3.row_dimensions[row].height = 24
    row += 1

# Column widths for legend
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 18
ws3.column_dimensions["E"].width = 18
ws3.column_dimensions["F"].width = 18

# ============================================================================
# SAVE
# ============================================================================

# ============================================================================
# SAVE ISSUE HISTORY
# ============================================================================

print(f"\n--- Saving issue history ---")
save_issue_history(issue_history)

print(f"\n--- Saving report ---")
try:
    wb.save(OUTPUT_PATH)
    print(f"  [OK] Report saved: {OUTPUT_PATH}")
    print(f"  [OK] File size: {OUTPUT_PATH.stat().st_size / 1024:.1f} KB")
except Exception as e:
    print(f"  [ERROR] Failed to save: {e}")
    sys.exit(1)

# ============================================================================
# SUMMARY STATS
# ============================================================================

print("\n" + "=" * 70)
print("  LISTING HEALTH REPORT - SUMMARY")
print("=" * 70)
print(f"  Report Date      : {REPORT_DATE}")
print(f"  ASINs Analyzed   : {len(asin_metrics)}")
print(f"  ---")
print(f"  Issues Summary:")
print(f"    NEW            : {status_counts.get('NEW', 0)}")
print(f"    REMINDER       : {status_counts.get('REMINDER', 0)}")
print(f"    ESCALATED      : {status_counts.get('ESCALATED', 0)}")
print(f"    RESOLVED       : {status_counts.get('RESOLVED', 0)}")
print(f"    SUPPRESSED     : {suppressed_count} (not shown)")
print(f"    In Report      : {len(issues_for_report)}")
print(f"  ---")
print(f"  Priority Breakdown:")
print(f"    HIGH           : {high_count}")
print(f"    MEDIUM         : {med_count}")
print(f"    LOW            : {low_count}")
print(f"  ---")
print(f"  Score Card       : {len(scorecards)} products")
print(f"    Green (75-100) : {green_count}")
print(f"    Blue  (50-74)  : {blue_count}")
print(f"    Yellow (25-49) : {yellow_count}")
print(f"    Red   (0-24)   : {red_count}")
print(f"  ---")
print(f"  Account Avg CTR  : {account_avg_ctr:.2f}%")
print(f"  Account Avg CVR  : {account_avg_cvr:.2f}%")
print(f"  Account Avg ACOS : {account_avg_acos:.2f}%")
print(f"  ---")
print(f"  History File     : {HISTORY_FILE.name}")
print(f"  Output File      : {OUTPUT_PATH.name}")
print("=" * 70)
print("  DONE!")
